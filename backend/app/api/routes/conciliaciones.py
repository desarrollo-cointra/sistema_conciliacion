from datetime import date
from io import BytesIO
import json
import zipfile
from decimal import Decimal, ROUND_HALF_UP

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user, is_cointra_admin
from app.core.config import settings
from app.db.session import get_db
from app.models.comentario import Comentario
from app.models.conciliacion import Conciliacion
from app.models.conciliacion_item import ConciliacionItem
from app.models.enums import ItemEstado, ItemTipo, UserRole
from app.models.factura_archivo_cliente import FacturaArchivoCliente
from app.models.historial_cambio import HistorialCambio
from app.models.operacion import Operacion
from app.models.usuario import Usuario
from app.models.usuario_operacion import usuario_operaciones_asignadas
from app.models.servicio import Servicio
from app.models.vehiculo import Vehiculo
from app.models.viaje import Viaje
from app.schemas.historial import HistorialCambioOut, ResumenFinancieroOut
from app.schemas.conciliacion import (
    ClienteItemDecision,
    ComentarioCreate,
    ComentarioOut,
    ConciliacionCreate,
    ConciliacionUpdate,
    ConciliacionItemCreate,
    ConciliacionItemOut,
    ConciliacionItemPatch,
    ConciliacionItemUpdateEstado,
    LiquidacionContratoFijoCreate,
    ConciliacionOut,
    ConciliacionWorkflowAction,
    ConciliacionUpdateEstado,
)
from app.services.pricing import apply_rentabilidad
from app.services.audit import log_change
from app.services.notifications import create_internal_notifications, send_manual_email
from app.services.avansat_cache import resolve_avansat_from_cache_only
from app.services.visibility import sanitize_item_for_role
from app.schemas.viaje import AdjuntarViajesRequest, ViajeOut

router = APIRouter(prefix="/conciliaciones", tags=["conciliaciones"])

TRANSPORTE_SERVICE_CODES = {"VIAJE", "VIAJE_ADICIONAL"}


def _login_url() -> str:
    return f"{settings.frontend_url.rstrip('/')}/login"


def _estado_conciliacion_viaje(viaje: Viaje):
    if not viaje.conciliacion:
        return None
    return getattr(viaje.conciliacion.estado, "value", viaje.conciliacion.estado)


def _should_mark_conciliado(estado: object) -> bool:
    estado_valor = getattr(estado, "value", estado)
    return str(estado_valor) in {"APROBADA", "CERRADA"}


def _default_viaje_item_financials(viaje: Viaje) -> tuple[float | None, float | None, float]:
    default_pct = 10.0
    tarifa_tercero = float(viaje.tarifa_tercero) if viaje.tarifa_tercero is not None else None

    if tarifa_tercero is not None:
        tarifa_cliente = tarifa_tercero / (1 - default_pct / 100)
        return tarifa_tercero, tarifa_cliente, default_pct

    tarifa_cliente = float(viaje.tarifa_cliente) if viaje.tarifa_cliente is not None else None
    if tarifa_cliente is not None:
        tarifa_tercero = tarifa_cliente * (1 - default_pct / 100)

    return tarifa_tercero, tarifa_cliente, default_pct


def _sync_viajes_conciliado_por_estado(db: Session, conciliacion_id: int, estado: object) -> None:
    estado_valor = getattr(estado, "value", estado)
    conciliado = _should_mark_conciliado(estado_valor)
    viajes = db.query(Viaje).filter(Viaje.conciliacion_id == conciliacion_id).all()
    for viaje in viajes:
        viaje.conciliado = conciliado
        viaje.estado_conciliacion = str(estado_valor)


def _repair_missing_viaje_items(db: Session, conc: Conciliacion, user_id: int) -> bool:
    """Repara inconsistencias historicas: viaje vinculado a conciliacion sin item VIAJE."""
    changed = False
    linked_viajes = db.query(Viaje).filter(Viaje.conciliacion_id == conc.id).all()
    existing_viaje_ids = {
        row[0]
        for row in db.query(ConciliacionItem.viaje_id)
        .filter(
            ConciliacionItem.conciliacion_id == conc.id,
            ConciliacionItem.tipo == ItemTipo.VIAJE,
            ConciliacionItem.viaje_id.is_not(None),
        )
        .all()
    }

    for viaje in linked_viajes:
        if viaje.id in existing_viaje_ids:
            continue

        tarifa_tercero, tarifa_cliente, rentabilidad = _default_viaje_item_financials(viaje)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            viaje_id=viaje.id,
            tipo=ItemTipo.VIAJE,
            fecha_servicio=viaje.fecha_servicio,
            origen=viaje.origen,
            destino=viaje.destino,
            placa=viaje.placa,
            conductor=viaje.conductor,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=tarifa_cliente,
            rentabilidad=rentabilidad,
            manifiesto_numero=viaje.manifiesto_numero,
            remesa=None,
            descripcion=viaje.descripcion,
            created_by=user_id,
            cargado_por=viaje.cargado_por,
        )
        db.add(item)
        log_change(
            db,
            usuario_id=user_id,
            conciliacion_id=conc.id,
            campo="reparacion_item_viaje",
            valor_nuevo=f"viaje_id={viaje.id}",
        )
        changed = True

    return changed


def _existing_item_viaje_ids(db: Session) -> set[int]:
    return {
        row[0]
        for row in db.query(ConciliacionItem.viaje_id)
        .filter(
            ConciliacionItem.tipo == ItemTipo.VIAJE,
            ConciliacionItem.viaje_id.is_not(None),
        )
        .all()
    }


def _validate_user_access_operacion(user: Usuario, operacion: Operacion) -> None:
    if user.rol == UserRole.CLIENTE:
        is_assigned = any(op.id == operacion.id for op in user.operaciones_asignadas)
        if not is_assigned:
            raise HTTPException(status_code=403, detail="Operacion no disponible para este cliente")
    if user.rol == UserRole.TERCERO and user.tercero_id != operacion.tercero_id:
        raise HTTPException(status_code=403, detail="Operacion no disponible para este tercero")


def _ensure_user_can_access_conciliacion(user: Usuario, conc: Conciliacion) -> None:
    if not conc.activo and not is_cointra_admin(user):
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    estado = getattr(conc.estado, "value", conc.estado)
    if user.rol == UserRole.CLIENTE and str(estado) == "BORRADOR":
        raise HTTPException(status_code=403, detail="La conciliacion aun no ha sido enviada a revision")


def _ensure_cointra_admin(user: Usuario) -> None:
    if not is_cointra_admin(user):
        raise HTTPException(status_code=403, detail="Solo COINTRA_ADMIN puede editar o inactivar conciliaciones")


def _parse_target_emails(raw_value: str | None, recipients: list[Usuario]) -> list[str]:
    provided: list[str] = []
    if raw_value:
        normalized = raw_value.replace(";", ",")
        provided = [email.strip() for email in normalized.split(",") if email and email.strip()]

    target_emails = provided or [u.email for u in recipients if u.email]
    return list(dict.fromkeys(target_emails))


def _users_matching_emails(recipients: list[Usuario], target_emails: list[str]) -> list[Usuario]:
    email_set = {email.strip().lower() for email in target_emails if email and email.strip()}
    if not email_set:
        return []
    matched: list[Usuario] = []
    seen_ids: set[int] = set()
    for user in recipients:
        if not user.email:
            continue
        if user.id in seen_ids:
            continue
        if user.email.strip().lower() in email_set:
            matched.append(user)
            seen_ids.add(user.id)
    return matched


def _sender_signature(user: Usuario) -> str:
    if user.email:
        return f"{user.nombre} <{user.email}>"
    return user.nombre


def _find_last_review_sender(db: Session, conciliacion_id: int) -> Usuario | None:
    last_sender_log = (
        db.query(HistorialCambio)
        .filter(
            HistorialCambio.conciliacion_id == conciliacion_id,
            HistorialCambio.campo == "enviar_revision",
        )
        .order_by(HistorialCambio.id.desc())
        .first()
    )
    if not last_sender_log:
        return None
    sender = db.get(Usuario, last_sender_log.usuario_id)
    if not sender or not sender.activo:
        return None
    return sender


def _resolve_recipients(db: Session, operacion: Operacion, roles: list[UserRole]) -> list[Usuario]:
    recipients: list[Usuario] = []
    for role in roles:
        query = db.query(Usuario).filter(Usuario.activo.is_(True), Usuario.rol == role)
        if role == UserRole.CLIENTE:
            query = query.join(
                usuario_operaciones_asignadas,
                usuario_operaciones_asignadas.c.usuario_id == Usuario.id,
            ).filter(
                Usuario.cliente_id == operacion.cliente_id,
                usuario_operaciones_asignadas.c.operacion_id == operacion.id,
            )
            recipients.extend(query.order_by(Usuario.id.asc()).all())
            continue
        elif role == UserRole.TERCERO:
            query = query.filter(Usuario.tercero_id == operacion.tercero_id)
        user = query.order_by(Usuario.id.asc()).first()
        if user:
            recipients.append(user)
    # Dedup por usuario
    uniq: dict[int, Usuario] = {u.id: u for u in recipients}
    return list(uniq.values())


def _display_estado(conc: Conciliacion) -> str:
    estado = str(getattr(conc.estado, "value", conc.estado))
    if estado == "CERRADA" and conc.factura_cliente_enviada:
        return "FACTURADO"
    if estado == "APROBADA" and conc.enviada_facturacion:
        return "ENVIADA_A_FACTURAR"
    return estado


def _mark_borrador_dirty(conc: Conciliacion) -> None:
    estado = str(getattr(conc.estado, "value", conc.estado))
    if estado == "BORRADOR":
        conc.borrador_guardado = False


def _next_liquidacion_id(db: Session, conciliacion_id: int) -> int:
    items = (
        db.query(ConciliacionItem)
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .all()
    )
    max_id = 0
    for item in items:
        meta = _extract_liquidacion_metadata(item)
        if not meta:
            continue
        liq_id = meta.get("liquidacion_contrato_fijo_id")
        if isinstance(liq_id, int) and liq_id > max_id:
            max_id = liq_id
    return max_id + 1


def _liquidacion_exists(db: Session, conciliacion_id: int, liquidacion_id: int) -> bool:
    if liquidacion_id <= 0:
        return False
    items = (
        db.query(ConciliacionItem)
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .all()
    )
    for item in items:
        meta = _extract_liquidacion_metadata(item)
        if not meta:
            continue
        if meta.get("liquidacion_contrato_fijo_id") == liquidacion_id:
            return True
    return False


def _find_last_status_actor(db: Session, conc: Conciliacion) -> tuple[str | None, str | None]:
    estado = _display_estado(conc)
    logs = (
        db.query(HistorialCambio)
        .filter(HistorialCambio.conciliacion_id == conc.id)
        .order_by(HistorialCambio.id.desc())
        .limit(150)
        .all()
    )

    def matches(log: HistorialCambio) -> bool:
        campo = (log.campo or "").strip()
        nuevo = str(log.valor_nuevo or "").strip().upper()

        if estado == "BORRADOR":
            return campo in {"devolucion_cliente", "conciliacion_creada"} or (
                campo == "estado_conciliacion" and nuevo == "BORRADOR"
            )
        if estado == "EN_REVISION":
            return campo == "enviar_revision" or (
                campo == "estado_conciliacion" and nuevo == "EN_REVISION"
            )
        if estado == "APROBADA":
            return campo == "aprobacion_cliente" or (
                campo == "estado_conciliacion" and nuevo == "APROBADA"
            )
        if estado == "ENVIADA_A_FACTURAR":
            return campo == "envio_facturacion"
        if estado == "CERRADA":
            return campo == "cierre_conciliacion" or (
                campo == "estado_conciliacion" and nuevo == "CERRADA"
            )
        if estado == "FACTURADO":
            return campo == "envio_factura_cliente" or (
                campo == "estado_conciliacion" and nuevo == "CERRADA"
            )
        return campo == "estado_conciliacion" and nuevo == estado

    for log in logs:
        if not matches(log):
            continue
        actor = db.get(Usuario, log.usuario_id)
        if actor:
            return actor.nombre, actor.email

    creator_name = conc.creador.nombre if conc.creador else None
    creator_email = conc.creador.email if conc.creador else None
    return creator_name, creator_email


def _build_conciliacion_totals_map(db: Session, conciliacion_ids: list[int]) -> dict[int, tuple[float, float]]:
    if not conciliacion_ids:
        return {}
    items = (
        db.query(ConciliacionItem)
        .filter(ConciliacionItem.conciliacion_id.in_(conciliacion_ids))
        .all()
    )

    totals_map: dict[int, tuple[float, float]] = {}
    for item in items:
        cid = int(item.conciliacion_id)
        # Bloque 1 (liquidación contrato fijo) es solo referencia, no se suma
        if _extract_liquidacion_metadata(item):
            continue
        current_cliente, current_tercero = totals_map.get(cid, (0.0, 0.0))
        totals_map[cid] = (
            current_cliente + float(item.tarifa_cliente or 0),
            current_tercero + float(item.tarifa_tercero or 0),
        )
    return totals_map


def _build_conciliacion_estado_timestamps(db: Session, conciliacion_id: int, created_at: object) -> dict[str, object | None]:
    logs = (
        db.query(HistorialCambio)
        .filter(HistorialCambio.conciliacion_id == conciliacion_id)
        .order_by(HistorialCambio.id.asc())
        .all()
    )
    timestamps: dict[str, object | None] = {
        "fecha_creacion": created_at,
        "fecha_envio_revision": None,
        "fecha_aprobacion": None,
        "fecha_rechazo": None,
        "fecha_envio_facturacion": None,
        "fecha_facturado": None,
    }
    for log in logs:
        campo = (log.campo or "").strip()
        if campo == "enviar_revision" and timestamps["fecha_envio_revision"] is None:
            timestamps["fecha_envio_revision"] = log.fecha
        elif campo == "aprobacion_cliente" and timestamps["fecha_aprobacion"] is None:
            timestamps["fecha_aprobacion"] = log.fecha
        elif campo == "devolucion_cliente" and timestamps["fecha_rechazo"] is None:
            timestamps["fecha_rechazo"] = log.fecha
        elif campo == "envio_facturacion" and timestamps["fecha_envio_facturacion"] is None:
            timestamps["fecha_envio_facturacion"] = log.fecha
        elif campo == "envio_factura_cliente" and timestamps["fecha_facturado"] is None:
            timestamps["fecha_facturado"] = log.fecha
    return timestamps


def _enrich_conciliacion(
    db: Session,
    conc: Conciliacion,
    user: Usuario,
    totals_map: dict[int, tuple[float, float]] | None = None,
) -> dict:
    """Convierte una Conciliacion ORM en dict con campos de creador, cliente y tercero."""
    base = ConciliacionOut.model_validate(conc).model_dump()
    base["creador_nombre"] = conc.creador.nombre if conc.creador else None
    operacion = conc.operacion
    base["cliente_nombre"] = operacion.cliente.nombre if operacion and operacion.cliente else None
    base["tercero_nombre"] = operacion.tercero.nombre if operacion and operacion.tercero else None
    estado_actor_nombre, estado_actor_email = _find_last_status_actor(db, conc)
    base["estado_actualizado_por_nombre"] = estado_actor_nombre
    base["estado_actualizado_por_email"] = estado_actor_email
    totals = (totals_map or {}).get(conc.id, (0.0, 0.0))
    valor_cliente, valor_tercero = totals
    if user.rol == UserRole.CLIENTE:
        base["valor_cliente"] = valor_cliente
        base["valor_tercero"] = None
    elif user.rol == UserRole.TERCERO:
        base["valor_cliente"] = None
        base["valor_tercero"] = valor_tercero
    else:
        base["valor_cliente"] = valor_cliente
        base["valor_tercero"] = valor_tercero
    timestamps = _build_conciliacion_estado_timestamps(db, conc.id, conc.created_at)
    base["fecha_creacion"] = timestamps["fecha_creacion"]
    base["fecha_envio_revision"] = timestamps["fecha_envio_revision"]
    base["fecha_aprobacion"] = timestamps["fecha_aprobacion"]
    base["fecha_rechazo"] = timestamps["fecha_rechazo"]
    base["fecha_envio_facturacion"] = timestamps["fecha_envio_facturacion"]
    base["fecha_facturado"] = timestamps["fecha_facturado"]
    return base


def _as_float(value: object) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def _split_total_evenly(total: float, count: int) -> list[float]:
    if count <= 0:
        return []
    total_dec = Decimal(str(total))
    if count == 1:
        return [float(total_dec)]

    base = (total_dec / Decimal(count)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    distributed: list[Decimal] = [base for _ in range(count - 1)]
    assigned = sum(distributed, Decimal("0.00"))
    distributed.append(total_dec - assigned)
    return [float(value) for value in distributed]


def _build_liquidacion_metadata(
    liquidacion_id: int,
    periodo_inicio: date,
    periodo_fin: date,
) -> str:
    payload = {
        "kind": "LIQUIDACION_CONTRATO_FIJO",
        "liquidacion_id": liquidacion_id,
        "periodo_inicio": str(periodo_inicio),
        "periodo_fin": str(periodo_fin),
    }
    return json.dumps(payload, ensure_ascii=True)


def _extract_liquidacion_metadata(item: ConciliacionItem) -> dict | None:
    if item.tipo != ItemTipo.OTRO:
        return None
    raw = (item.descripcion or "").strip()
    if not raw:
        return None

    try:
        payload = json.loads(raw)
    except Exception:
        return None

    if payload.get("kind") != "LIQUIDACION_CONTRATO_FIJO":
        return None

    try:
        periodo_inicio = date.fromisoformat(str(payload.get("periodo_inicio") or ""))
        periodo_fin = date.fromisoformat(str(payload.get("periodo_fin") or ""))
    except Exception:
        return None

    return {
        "liquidacion_contrato_fijo": True,
        "liquidacion_contrato_fijo_id": payload.get("liquidacion_id"),
        "liquidacion_periodo_inicio": periodo_inicio,
        "liquidacion_periodo_fin": periodo_fin,
    }


def _normalize_manifiesto_for_lookup(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    # Corrige casos donde el manifiesto llega como numero decimal de Excel, p.e. "522318.0"
    if raw.endswith(".0"):
        integer_part = raw[:-2]
        if integer_part.isdigit():
            return integer_part
    return raw


def _normalize_placa_for_compare(value: object) -> str:
    raw = str(value or "").strip().upper()
    return "".join(ch for ch in raw if ch.isalnum())


def _item_servicio_codigo(item: ConciliacionItem) -> str:
    viaje = getattr(item, "viaje", None)
    servicio = getattr(viaje, "servicio", None)
    raw_codigo = getattr(servicio, "codigo", "")
    return str(raw_codigo or "").strip().upper()


def _is_transport_item(item: ConciliacionItem) -> bool:
    codigo = _item_servicio_codigo(item)
    return codigo in TRANSPORTE_SERVICE_CODES


def _validate_transport_item_manifest_or_raise(db: Session, item: ConciliacionItem) -> None:
    if not _is_transport_item(item):
        return

    manifiesto = _normalize_manifiesto_for_lookup(item.manifiesto_numero)
    if not manifiesto:
        servicio_codigo = _item_servicio_codigo(item) or "VIAJE"
        raise HTTPException(
            status_code=400,
            detail=(
                f"El servicio {servicio_codigo} requiere manifiesto. "
                "Asocia un manifiesto antes de continuar."
            ),
        )

    avansat = _fetch_avansat_with_fallback(db, manifiesto)
    if not avansat:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El manifiesto {manifiesto} no existe en la cache de Avansat. "
                "Sincroniza Avansat o corrige el numero del manifiesto."
            ),
        )

    placa_servicio = _normalize_placa_for_compare(item.placa)
    if not placa_servicio:
        raise HTTPException(
            status_code=400,
            detail="El servicio de transporte no tiene placa registrada para validar el manifiesto.",
        )

    placa_avansat = _normalize_placa_for_compare(avansat.get("placa_vehiculo") or "")
    if not placa_avansat:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El manifiesto {manifiesto} no tiene placa valida en Avansat. "
                "Corrige el manifiesto o sincroniza la fuente."
            ),
        )

    if placa_avansat != placa_servicio:
        raise HTTPException(
            status_code=400,
            detail=(
                f"La placa del manifiesto {manifiesto} ({placa_avansat}) no coincide con la placa del servicio ({placa_servicio})."
            ),
        )

    # Un manifiesto solo puede estar asociado a un único ítem/viaje.
    existing = (
        db.query(ConciliacionItem.id)
        .filter(
            ConciliacionItem.manifiesto_numero == manifiesto,
            ConciliacionItem.id != item.id,
        )
        .first()
    )
    if existing:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El manifiesto {manifiesto} ya esta asociado a otro viaje "
                f"(item #{existing.id}). Un manifiesto solo puede asociarse a un viaje."
            ),
        )


def _validate_transport_items_manifests_or_raise(
    db: Session,
    items: list[ConciliacionItem],
    action_label: str,
) -> None:
    errors: list[str] = []
    for item in items:
        if not _is_transport_item(item):
            continue
        try:
            _validate_transport_item_manifest_or_raise(db, item)
        except HTTPException as exc:
            detail = exc.detail if isinstance(exc.detail, str) else "Error de validacion"
            item_ref = f"viaje #{item.viaje_id}" if item.viaje_id else f"item #{item.id}"
            errors.append(f"{item_ref}: {detail}")

    if errors:
        preview = "\n".join(errors[:10])
        suffix = "" if len(errors) <= 10 else f"\n... y {len(errors) - 10} errores mas"
        raise HTTPException(
            status_code=400,
            detail=(
                f"No se puede {action_label}. Los servicios de transporte (VIAJE/VIAJE_ADICIONAL) deben tener manifiesto valido y placa coincidente.\n"
                f"{preview}{suffix}"
            ),
        )


def _fetch_avansat_with_fallback(
    db: Session,
    manifiesto: str,
    prefetched: dict[str, dict] | None = None,
) -> dict:
    if not manifiesto:
        return {}
    attempts = [manifiesto, manifiesto.lstrip("0")]
    seen: set[str] = set()
    for candidate in attempts:
        candidate = (candidate or "").strip()
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        if prefetched:
            prefetched_data = prefetched.get(candidate) or {}
            if prefetched_data:
                return prefetched_data
        resolved, _ = resolve_avansat_from_cache_only(db, [candidate])
        data = resolved.get(candidate) or {}
        if data:
            return data
    return {}


def _prefetch_avansat_for_manifest_numbers_or_raise(
    db: Session,
    manifest_numbers: list[str],
) -> dict[str, dict]:
    unique_manifiestos = list(
        dict.fromkeys(
            [
                _normalize_manifiesto_for_lookup(number)
                for number in manifest_numbers
                if _normalize_manifiesto_for_lookup(number)
            ]
        )
    )
    if not unique_manifiestos:
        return {}

    prefetched, missing = resolve_avansat_from_cache_only(db, unique_manifiestos)
    if missing:
        total = len(unique_manifiestos)
        missing_count = len(missing)
        resolved_count = total - missing_count
        missing_list = ", ".join(missing[:10])
        suffix = "" if len(missing) <= 10 else f" y {len(missing) - 10} mas"
        raise HTTPException(
            status_code=502,
            detail=(
                "La cache interna de Avansat aun no tiene todos los manifiestos requeridos para generar el Excel. "
                f"Disponibles: {resolved_count}/{total}. "
                f"Faltantes ({missing_count}): {missing_list}{suffix}. "
                "Espera la siguiente sincronizacion automatica (cada 30 minutos) o ejecuta una sincronizacion manual desde Consulta Avansat."
            ),
        )

    return prefetched


def _prepare_facturacion_rows(
    db: Session,
    items: list[ConciliacionItem],
    avansat_prefetched: dict[str, dict] | None = None,
) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    missing_manifiestos: list[str] = []

    for item in items:
        viaje_ref = f"viaje #{item.viaje_id}" if item.viaje_id else f"item #{item.id}"
        manifiesto = _normalize_manifiesto_for_lookup(item.manifiesto_numero)
        if not manifiesto:
            missing_manifiestos.append(f"{viaje_ref} (sin manifiesto)")
            continue

        avansat = _fetch_avansat_with_fallback(db, manifiesto, avansat_prefetched)
        if not avansat:
            missing_manifiestos.append(f"{viaje_ref} (manifiesto {manifiesto} sin datos en Avansat)")
            continue

        precio_cliente = _as_float(item.tarifa_cliente)
        precio_tercero = _as_float(item.tarifa_tercero)
        rentabilidad = _as_float(item.rentabilidad)
        ganancia = precio_cliente - precio_tercero

        rows.append(
            {
                "manifiesto": manifiesto,
                "fecha_emision": avansat.get("fecha_emision") or "",
                "producto": avansat.get("producto") or "",
                "placa_vehiculo": avansat.get("placa_vehiculo") or (item.placa or ""),
                "trayler": avansat.get("trayler") or "",
                "remesa": avansat.get("remesa") or (item.remesa or ""),
                "ciudad_origen": avansat.get("ciudad_origen") or (item.origen or ""),
                "ciudad_destino": avansat.get("ciudad_destino") or (item.destino or ""),
                "precio_cliente": precio_cliente,
                "precio_tercero": precio_tercero,
                "rentabilidad": rentabilidad,
                "ganancia": ganancia,
            }
        )

    return rows, sorted(set(missing_manifiestos))


def _build_facturacion_excel(conc: Conciliacion, rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturacion"

    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="FFF200")
    header_font = Font(bold=True, color="1F2937")
    total_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cop_format = '"$" #,##0'

    ws.append(
        [
            "Manifiesto",
            "Fecha Manifiesto",
            "Producto",
            "Placa",
            "Remolque",
            "Remesa",
            "Origen",
            "Destino",
            "Precio Cliente",
            "Precio Tercero",
            "Rentabilidad",
            "Ganancia Cointra",
        ]
    )

    for idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        if idx in (9, 10, 11, 12):
            cell.fill = total_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    total_precio_cliente = 0.0
    total_precio_tercero = 0.0
    total_ganancia = 0.0

    for row in rows:
        manifiesto = str(row.get("manifiesto") or "").strip()
        precio_cliente = _as_float(row.get("precio_cliente"))
        precio_tercero = _as_float(row.get("precio_tercero"))
        rentabilidad = _as_float(row.get("rentabilidad"))
        ganancia = _as_float(row.get("ganancia"))

        total_precio_cliente += precio_cliente
        total_precio_tercero += precio_tercero
        total_ganancia += ganancia

        ws.append(
            [
                manifiesto,
                row.get("fecha_emision") or "",
                row.get("producto") or "",
                row.get("placa_vehiculo") or "",
                row.get("trayler") or "",
                row.get("remesa") or "",
                row.get("ciudad_origen") or "",
                row.get("ciudad_destino") or "",
                precio_cliente,
                precio_tercero,
                rentabilidad,
                ganancia,
            ]
        )

        row_idx = ws.max_row
        for col_idx in range(1, 13):
            ws.cell(row=row_idx, column=col_idx).border = border

        for col_idx in (9, 10, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = cop_format

        ws.cell(row=row_idx, column=11).number_format = '#,##0.##" %"'

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=8, value="TOTAL")
    ws.cell(row=total_row, column=9, value=total_precio_cliente)
    ws.cell(row=total_row, column=10, value=total_precio_tercero)
    ws.cell(row=total_row, column=12, value=total_ganancia)

    for col_idx in (8, 9, 10, 12):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if col_idx in (9, 10, 12):
            cell.number_format = cop_format

    widths = {
        1: 14,
        2: 18,
        3: 24,
        4: 12,
        5: 12,
        6: 12,
        7: 16,
        8: 16,
        9: 18,
        10: 18,
        11: 14,
        12: 18,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_conciliacion_excel_legacy(
    conc: Conciliacion,
    items: list[ConciliacionItem],
    user_role: UserRole,
    tipo_vehiculo_by_placa: dict[str, str],
    avansat_prefetched: dict[str, dict] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "resumen"

    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    header_font = Font(bold=True, color="1F2937")
    section_fill = PatternFill(fill_type="solid", fgColor="FDE68A")
    section_font = Font(bold=True, color="111827")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cop_format = '"$" #,##0'
    pct_format = '#,##0.##" %"'

    show_tarifa_tercero = user_role != UserRole.CLIENTE
    show_tarifa_cliente = user_role != UserRole.TERCERO
    show_cointra_financials = user_role == UserRole.COINTRA

    liquidacion_items: list[tuple[ConciliacionItem, dict]] = []
    additional_items: list[tuple[ConciliacionItem, dict | None]] = []
    for item in items:
        liq_meta = _extract_liquidacion_metadata(item)
        if liq_meta:
            liquidacion_items.append((item, liq_meta))
        else:
            service_code = str(item.servicio_codigo or "").strip().upper()
            if service_code == "DISPONIBILIDAD":
                continue
            additional_items.append((item, None))

    estado_display = _display_estado(conc)
    estado_label = estado_display.replace("_", " ")
    estado_styles = {
        "BORRADOR": ("ECFDF5", "065F46"),
        "EN_REVISION": ("FFFBEB", "92400E"),
        "APROBADA": ("F0FDFA", "115E59"),
        "ENVIADA_A_FACTURAR": ("F0F9FF", "075985"),
    }
    estado_fill_color, estado_font_color = estado_styles.get(estado_display, ("F7FEE7", "3F6212"))

    current_row = 1
    ws.cell(row=current_row, column=1, value=f"Conciliacion #{conc.id} - {conc.nombre}")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="111827")
    estado_cell = ws.cell(row=current_row, column=3, value=f"ESTADO: {estado_label}")
    estado_cell.fill = PatternFill(fill_type="solid", fgColor=estado_fill_color)
    estado_cell.font = Font(bold=True, color=estado_font_color)
    estado_cell.alignment = Alignment(horizontal="center", vertical="center")
    estado_cell.border = border
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}")
    ws.cell(row=current_row, column=1).font = Font(size=10, color="374151")
    current_row += 2

    # Bloque superior: liquidacion contrato fijo
    ws.cell(row=current_row, column=1, value="LIQUIDACION CONTRATO FIJO")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="111827")
    current_row += 1

    top_headers = ["Placa", "Tipo Vehiculo"]
    if show_tarifa_cliente:
        top_headers.append("Valor Cliente")
    if show_tarifa_tercero:
        top_headers.append("Valor Tercero")
    if show_cointra_financials:
        top_headers.extend(["Rentabilidad", "Ganancia Cointra"])

    top_col_idx = {header: idx for idx, header in enumerate(top_headers, start=1)}
    for idx, header in enumerate(top_headers, start=1):
        cell = ws.cell(row=current_row, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    current_row += 1

    top_start_row = current_row
    liquidacion_items_sorted = sorted(
        liquidacion_items,
        key=lambda pair: (
            str(pair[0].placa or "").upper(),
            pair[0].fecha_servicio,
            pair[0].id,
        ),
    )

    for item, liq_meta in liquidacion_items_sorted:
        placa = (item.placa or "").upper()
        tipo_vehiculo = tipo_vehiculo_by_placa.get(placa, "")

        tarifa_tercero = _as_float(item.tarifa_tercero)
        tarifa_cliente = _as_float(item.tarifa_cliente)
        ganancia = tarifa_cliente - tarifa_tercero
        rentabilidad = _as_float(item.rentabilidad)

        ws.cell(row=current_row, column=top_col_idx["Placa"], value=placa)
        ws.cell(row=current_row, column=top_col_idx["Tipo Vehiculo"], value=tipo_vehiculo)
        if "Valor Cliente" in top_col_idx:
            c = ws.cell(row=current_row, column=top_col_idx["Valor Cliente"], value=tarifa_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in top_col_idx:
            c = ws.cell(row=current_row, column=top_col_idx["Valor Tercero"], value=tarifa_tercero)
            c.number_format = cop_format
        if "Rentabilidad" in top_col_idx:
            c = ws.cell(row=current_row, column=top_col_idx["Rentabilidad"], value=rentabilidad)
            c.number_format = pct_format
        if "Ganancia Cointra" in top_col_idx:
            c = ws.cell(row=current_row, column=top_col_idx["Ganancia Cointra"], value=ganancia)
            c.number_format = cop_format

        for col_idx in range(1, len(top_headers) + 1):
            ws.cell(row=current_row, column=col_idx).border = border
        current_row += 1

    if current_row == top_start_row:
        ws.cell(row=current_row, column=1, value="(sin registros de contrato fijo)")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="6B7280")
        current_row += 1

    top_total_tercero = sum(_as_float(item.tarifa_tercero) for item, _ in liquidacion_items_sorted)
    top_total_cliente = sum(_as_float(item.tarifa_cliente) for item, _ in liquidacion_items_sorted)
    top_ganancia = top_total_cliente - top_total_tercero
    top_rentabilidad = (top_ganancia / top_total_cliente * 100) if top_total_cliente > 0 else 0.0

    current_row += 1
    ws.cell(row=current_row, column=2, value="TOTAL CONTRATO FIJO")
    if "Valor Cliente" in top_col_idx:
        c = ws.cell(row=current_row, column=top_col_idx["Valor Cliente"], value=top_total_cliente)
        c.number_format = cop_format
    if "Valor Tercero" in top_col_idx:
        c = ws.cell(row=current_row, column=top_col_idx["Valor Tercero"], value=top_total_tercero)
        c.number_format = cop_format
    if "Rentabilidad" in top_col_idx:
        c = ws.cell(row=current_row, column=top_col_idx["Rentabilidad"], value=top_rentabilidad)
        c.number_format = pct_format
    if "Ganancia Cointra" in top_col_idx:
        c = ws.cell(row=current_row, column=top_col_idx["Ganancia Cointra"], value=top_ganancia)
        c.number_format = cop_format

    for col_idx in range(1, len(top_headers) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = section_fill
        cell.font = section_font
        cell.border = border

    # Bloque inferior: adicionales/otros servicios
    current_row += 2
    ws.cell(row=current_row, column=1, value="SERVICIOS")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="111827")
    current_row += 1

    bottom_headers = ["Placa", "Tipo Vehiculo", "Fecha", "Titulo Servicio", "Tipo Servicio"]
    if show_tarifa_cliente:
        bottom_headers.append("Valor Cliente")
    if show_tarifa_tercero:
        bottom_headers.append("Valor Tercero")    
    if show_cointra_financials:
        bottom_headers.extend(["Rentabilidad", "Ganancia Cointra"])
    bottom_headers.append("Observaciones")
    bottom_col_idx = {header: idx for idx, header in enumerate(bottom_headers, start=1)}

    for idx, header in enumerate(bottom_headers, start=1):
        cell = ws.cell(row=current_row, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    current_row += 1

    additional_sorted = sorted(
        additional_items,
        key=lambda pair: (
            str(pair[0].placa or "").upper(),
            pair[0].fecha_servicio,
            pair[0].id,
        ),
    )

    group_placa = None
    group_total_tercero = 0.0
    group_total_cliente = 0.0
    all_total_tercero = 0.0
    all_total_cliente = 0.0

    def _write_group_total(row_num: int, placa: str, total_tercero: float, total_cliente: float) -> int:
        if not placa:
            return row_num
        ganancia = total_cliente - total_tercero
        rentabilidad = (ganancia / total_cliente * 100) if total_cliente > 0 else 0.0
        ws.cell(row=row_num, column=2, value=f"TOTAL PLACA {placa}")
        if "Valor Cliente" in bottom_col_idx:
            c = ws.cell(row=row_num, column=bottom_col_idx["Valor Cliente"], value=total_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in bottom_col_idx:
            c = ws.cell(row=row_num, column=bottom_col_idx["Valor Tercero"], value=total_tercero)
            c.number_format = cop_format        
        if "Rentabilidad" in bottom_col_idx:
            c = ws.cell(row=row_num, column=bottom_col_idx["Rentabilidad"], value=rentabilidad)
            c.number_format = pct_format
        if "Ganancia Cointra" in bottom_col_idx:
            c = ws.cell(row=row_num, column=bottom_col_idx["Ganancia Cointra"], value=ganancia)
            c.number_format = cop_format

        for col_idx in range(1, len(bottom_headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = section_font
            cell.border = border
        return row_num + 1

    for item, _ in additional_sorted:
        placa = (item.placa or "").upper() or "SIN_PLACA"
        if group_placa is None:
            group_placa = placa
        elif placa != group_placa:
            current_row = _write_group_total(current_row, group_placa, group_total_tercero, group_total_cliente)
            group_placa = placa
            group_total_tercero = 0.0
            group_total_cliente = 0.0

        tipo_vehiculo = tipo_vehiculo_by_placa.get(placa, "")
        titulo_servicio = item.viaje.titulo if item.viaje else ""
        tipo_servicio = (
            (item.servicio_nombre or "").strip()
            or (item.servicio_codigo or "").strip()
            or str(getattr(item.tipo, "value", item.tipo))
        )
        tarifa_tercero = _as_float(item.tarifa_tercero)
        tarifa_cliente = _as_float(item.tarifa_cliente)
        ganancia = tarifa_cliente - tarifa_tercero
        rentabilidad = _as_float(item.rentabilidad)

        group_total_tercero += tarifa_tercero
        group_total_cliente += tarifa_cliente
        all_total_tercero += tarifa_tercero
        all_total_cliente += tarifa_cliente

        ws.cell(row=current_row, column=bottom_col_idx["Placa"], value=placa)
        ws.cell(row=current_row, column=bottom_col_idx["Tipo Vehiculo"], value=tipo_vehiculo)
        ws.cell(row=current_row, column=bottom_col_idx["Fecha"], value=str(item.fecha_servicio))
        ws.cell(row=current_row, column=bottom_col_idx["Titulo Servicio"], value=titulo_servicio)
        ws.cell(row=current_row, column=bottom_col_idx["Tipo Servicio"], value=tipo_servicio)
        if "Valor Cliente" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Valor Cliente"], value=tarifa_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Valor Tercero"], value=tarifa_tercero)
            c.number_format = cop_format        
        if "Rentabilidad" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Rentabilidad"], value=rentabilidad)
            c.number_format = pct_format
        if "Ganancia Cointra" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Ganancia Cointra"], value=ganancia)
            c.number_format = cop_format
        ws.cell(row=current_row, column=bottom_col_idx["Observaciones"], value=item.descripcion or "")

        for col_idx in range(1, len(bottom_headers) + 1):
            ws.cell(row=current_row, column=col_idx).border = border
        current_row += 1

    if not additional_sorted:
        ws.cell(row=current_row, column=1, value="(sin servicios adicionales)")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="6B7280")
        current_row += 1
    else:
        current_row = _write_group_total(current_row, group_placa or "", group_total_tercero, group_total_cliente)

        all_ganancia = all_total_cliente - all_total_tercero
        all_rentabilidad = (all_ganancia / all_total_cliente * 100) if all_total_cliente > 0 else 0.0
        ws.cell(row=current_row, column=2, value="TOTAL SERVICIOS")
        if "Valor Cliente" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Valor Cliente"], value=all_total_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Valor Tercero"], value=all_total_tercero)
            c.number_format = cop_format
        if "Rentabilidad" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Rentabilidad"], value=all_rentabilidad)
            c.number_format = pct_format
        if "Ganancia Cointra" in bottom_col_idx:
            c = ws.cell(row=current_row, column=bottom_col_idx["Ganancia Cointra"], value=all_ganancia)
            c.number_format = cop_format
        for col_idx in range(1, len(bottom_headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = section_fill
            cell.font = section_font
            cell.border = border
        current_row += 2

    # Hoja 2: servicios (manifiestos del contexto conciliacion + datos Avansat)
    manifest_financials_servicios: dict[str, dict[str, float]] = {}
    servicios_totals_by_placa: dict[str, dict[str, float]] = {}
    for additional_item, _ in additional_items:
        manifest_key = _normalize_manifiesto_for_lookup(additional_item.manifiesto_numero)
        if manifest_key:
            bucket = manifest_financials_servicios.setdefault(
                manifest_key,
                {"valor_tercero": 0.0, "valor_cliente": 0.0},
            )
            bucket["valor_tercero"] += _as_float(additional_item.tarifa_tercero)
            bucket["valor_cliente"] += _as_float(additional_item.tarifa_cliente)

        placa_key = str(additional_item.placa or "").strip().upper()
        if placa_key:
            plate_bucket = servicios_totals_by_placa.setdefault(
                placa_key,
                {"valor_tercero": 0.0, "valor_cliente": 0.0},
            )
            plate_bucket["valor_tercero"] += _as_float(additional_item.tarifa_tercero)
            plate_bucket["valor_cliente"] += _as_float(additional_item.tarifa_cliente)

    ws_servicios = wb.create_sheet("Servicios")
    servicios_headers = [
        "Manifiesto",
        "Fecha Emision",
        "Placa Vehiculo",
        "Trayler",
        "Remesa",
        "Producto",
        "Ciudad Origen",
        "Ciudad Destino",
    ]
    if show_tarifa_cliente:
        servicios_headers.append("Valor Cliente")
    if show_tarifa_tercero:
        servicios_headers.append("Valor Tercero")
    if show_cointra_financials:
        servicios_headers.extend(["Rentabilidad", "Ganancia Cointra"])

    servicios_col_idx = {header: idx for idx, header in enumerate(servicios_headers, start=1)}
    ws_servicios.append(servicios_headers)
    for idx, cell in enumerate(ws_servicios[1], start=1):
        cell.fill = header_fill
        if servicios_headers[idx - 1] in {"Valor Tercero", "Valor Cliente", "Rentabilidad", "Ganancia Cointra"}:
            cell.fill = section_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    def _write_servicios_block_header(row_idx: int) -> None:
        for idx, header in enumerate(servicios_headers, start=1):
            cell = ws_servicios.cell(row=row_idx, column=idx, value=header)
            cell.fill = header_fill
            if header in {"Valor Tercero", "Valor Cliente", "Rentabilidad", "Ganancia Cointra"}:
                cell.fill = section_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    servicios_entries: list[dict[str, object]] = []
    manifests_sorted = sorted(manifest_financials_servicios.keys())
    avansat_servicios_prefetched = avansat_prefetched or {}

    for manifest_key in manifests_sorted:
        manifiesto = _normalize_manifiesto_for_lookup(manifest_key)
        avansat = avansat_servicios_prefetched.get(manifiesto) or {}
        remesas = avansat.get("remesas") if isinstance(avansat.get("remesas"), list) else []
        remesas_rows = [r for r in remesas if isinstance(r, dict)]
        if not remesas_rows:
            remesas_rows = [{"remesa": avansat.get("remesa") or "", "producto": avansat.get("producto") or ""}]

        financials = manifest_financials_servicios.get(manifiesto, {"valor_tercero": 0.0, "valor_cliente": 0.0})
        valor_tercero = _as_float(financials.get("valor_tercero"))
        valor_cliente = _as_float(financials.get("valor_cliente"))
        ganancia = valor_cliente - valor_tercero
        rentabilidad = (ganancia / valor_cliente * 100) if valor_cliente > 0 else 0.0
        placa = str(avansat.get("placa_vehiculo") or "").strip().upper() or "SIN_PLACA"

        servicios_entries.append(
            {
                "placa": placa,
                "manifiesto": manifiesto,
                "fecha_emision": avansat.get("fecha_emision") or "",
                "trayler": avansat.get("trayler") or "",
                "ciudad_origen": avansat.get("ciudad_origen") or "",
                "ciudad_destino": avansat.get("ciudad_destino") or "",
                "remesas": remesas_rows,
                "valor_tercero": valor_tercero,
                "valor_cliente": valor_cliente,
                "rentabilidad": rentabilidad,
                "ganancia": ganancia,
            }
        )

    servicios_entries.sort(key=lambda row: (str(row["placa"]), str(row["manifiesto"])))

    if not servicios_entries:
        ws_servicios.append(["Sin manifiestos asociados a conciliacion"])
        ws_servicios.cell(row=2, column=1).font = Font(italic=True, color="6B7280")
    else:
        current_row = 2
        total_general_tercero = 0.0
        total_general_cliente = 0.0
        placas = sorted({str(row["placa"]) for row in servicios_entries})

        for plate_index, placa in enumerate(placas):
            plate_rows = [row for row in servicios_entries if str(row["placa"]) == placa]
            total_placa_tercero = 0.0
            total_placa_cliente = 0.0

            servicios_totals = servicios_totals_by_placa.get(placa)
            if servicios_totals:
                total_placa_tercero = _as_float(servicios_totals.get("valor_tercero"))
                total_placa_cliente = _as_float(servicios_totals.get("valor_cliente"))
            else:
                # Fallback: si no hay registro de servicios para la placa, conserva suma por manifiestos.
                total_placa_tercero = sum(_as_float(row.get("valor_tercero")) for row in plate_rows)
                total_placa_cliente = sum(_as_float(row.get("valor_cliente")) for row in plate_rows)

            manifest_count = len(plate_rows)
            per_manifest_tercero = _split_total_evenly(total_placa_tercero, manifest_count)
            per_manifest_cliente = _split_total_evenly(total_placa_cliente, manifest_count)

            if plate_index > 0:
                row_num_spacer = current_row
                current_row += 1
                _write_servicios_block_header(current_row)
                current_row += 1

            for entry_idx, entry in enumerate(plate_rows):
                remesas_rows = entry["remesas"] if isinstance(entry["remesas"], list) else []
                manifest_valor_tercero = per_manifest_tercero[entry_idx] if entry_idx < len(per_manifest_tercero) else 0.0
                manifest_valor_cliente = per_manifest_cliente[entry_idx] if entry_idx < len(per_manifest_cliente) else 0.0
                manifest_ganancia = manifest_valor_cliente - manifest_valor_tercero
                manifest_rentabilidad = (
                    (manifest_ganancia / manifest_valor_cliente * 100) if manifest_valor_cliente > 0 else 0.0
                )
                first_row_for_manifest = True
                for remesa_row in remesas_rows:
                    row_values: dict[str, object] = {
                        "Manifiesto": entry["manifiesto"],
                        "Fecha Emision": entry["fecha_emision"],
                        "Placa Vehiculo": entry["placa"],
                        "Trayler": entry["trayler"],
                        "Remesa": str((remesa_row or {}).get("remesa") or "").strip(),
                        "Producto": str((remesa_row or {}).get("producto") or "").strip(),
                        "Ciudad Origen": entry["ciudad_origen"],
                        "Ciudad Destino": entry["ciudad_destino"],
                    }
                    if show_tarifa_tercero:
                        row_values["Valor Tercero"] = manifest_valor_tercero if first_row_for_manifest else None
                    if show_tarifa_cliente:
                        row_values["Valor Cliente"] = manifest_valor_cliente if first_row_for_manifest else None
                    if show_cointra_financials:
                        row_values["Rentabilidad"] = manifest_rentabilidad if first_row_for_manifest else None
                        row_values["Ganancia Cointra"] = manifest_ganancia if first_row_for_manifest else None

                    for header, col_idx in servicios_col_idx.items():
                        cell = ws_servicios.cell(row=current_row, column=col_idx, value=row_values.get(header, ""))
                        cell.border = border
                        if header in {"Valor Tercero", "Valor Cliente", "Ganancia Cointra"} and cell.value is not None:
                            cell.number_format = cop_format
                        if header == "Rentabilidad" and cell.value is not None:
                            cell.number_format = pct_format
                    current_row += 1
                    first_row_for_manifest = False

            total_general_tercero += total_placa_tercero
            total_general_cliente += total_placa_cliente

            ws_servicios.cell(row=current_row, column=servicios_col_idx["Ciudad Destino"], value=f"TOTAL PLACA {placa}")
            total_placa_cells_to_fill = [servicios_col_idx["Ciudad Destino"]]
            if "Valor Tercero" in servicios_col_idx:
                c = ws_servicios.cell(row=current_row, column=servicios_col_idx["Valor Tercero"], value=total_placa_tercero)
                c.number_format = cop_format
                total_placa_cells_to_fill.append(servicios_col_idx["Valor Tercero"])
            if "Valor Cliente" in servicios_col_idx:
                c = ws_servicios.cell(row=current_row, column=servicios_col_idx["Valor Cliente"], value=total_placa_cliente)
                c.number_format = cop_format
                total_placa_cells_to_fill.append(servicios_col_idx["Valor Cliente"])
            if "Rentabilidad" in servicios_col_idx:
                placa_ganancia = total_placa_cliente - total_placa_tercero
                placa_rentabilidad = (placa_ganancia / total_placa_cliente * 100) if total_placa_cliente > 0 else 0.0
                c = ws_servicios.cell(row=current_row, column=servicios_col_idx["Rentabilidad"], value=placa_rentabilidad)
                c.number_format = pct_format
                total_placa_cells_to_fill.append(servicios_col_idx["Rentabilidad"])
            if "Ganancia Cointra" in servicios_col_idx:
                c = ws_servicios.cell(
                    row=current_row,
                    column=servicios_col_idx["Ganancia Cointra"],
                    value=total_placa_cliente - total_placa_tercero,
                )
                c.number_format = cop_format
                total_placa_cells_to_fill.append(servicios_col_idx["Ganancia Cointra"])

            for col_idx in range(1, len(servicios_headers) + 1):
                cell = ws_servicios.cell(row=current_row, column=col_idx)
                cell.font = section_font
                cell.border = border
            for col_idx in total_placa_cells_to_fill:
                ws_servicios.cell(row=current_row, column=col_idx).fill = section_fill
            current_row += 1

        current_row += 2
        ws_servicios.cell(row=current_row, column=servicios_col_idx["Ciudad Destino"], value="TOTAL GENERAL")
        total_general_cells_to_fill = [servicios_col_idx["Ciudad Destino"]]
        if "Valor Tercero" in servicios_col_idx:
            c = ws_servicios.cell(row=current_row, column=servicios_col_idx["Valor Tercero"], value=total_general_tercero)
            c.number_format = cop_format
            total_general_cells_to_fill.append(servicios_col_idx["Valor Tercero"])
        if "Valor Cliente" in servicios_col_idx:
            c = ws_servicios.cell(row=current_row, column=servicios_col_idx["Valor Cliente"], value=total_general_cliente)
            c.number_format = cop_format
            total_general_cells_to_fill.append(servicios_col_idx["Valor Cliente"])
        if "Rentabilidad" in servicios_col_idx:
            total_ganancia = total_general_cliente - total_general_tercero
            total_rentabilidad = (total_ganancia / total_general_cliente * 100) if total_general_cliente > 0 else 0.0
            c = ws_servicios.cell(row=current_row, column=servicios_col_idx["Rentabilidad"], value=total_rentabilidad)
            c.number_format = pct_format
            total_general_cells_to_fill.append(servicios_col_idx["Rentabilidad"])
        if "Ganancia Cointra" in servicios_col_idx:
            c = ws_servicios.cell(
                row=current_row,
                column=servicios_col_idx["Ganancia Cointra"],
                value=total_general_cliente - total_general_tercero,
            )
            c.number_format = cop_format
            total_general_cells_to_fill.append(servicios_col_idx["Ganancia Cointra"])

        for col_idx in range(1, len(servicios_headers) + 1):
            cell = ws_servicios.cell(row=current_row, column=col_idx)
            cell.font = section_font
            cell.border = border
        for col_idx in total_general_cells_to_fill:
            ws_servicios.cell(row=current_row, column=col_idx).fill = section_fill

    servicios_widths = {
        "Manifiesto": 20,
        "Fecha Emision": 18,
        "Placa Vehiculo": 20,
        "Trayler": 18,
        "Remesa": 20,
        "Producto": 40,
        "Ciudad Origen": 24,
        "Ciudad Destino": 24,
        "Valor Tercero": 20,
        "Valor Cliente": 20,
        "Rentabilidad": 16,
        "Ganancia Cointra": 22,
    }
    for header, idx in servicios_col_idx.items():
        ws_servicios.column_dimensions[get_column_letter(idx)].width = servicios_widths.get(header, 18)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_conciliacion_excel(
    conc: Conciliacion,
    items: list[ConciliacionItem],
    user_role: UserRole,
    tipo_vehiculo_by_placa: dict[str, str],
    avansat_prefetched: dict[str, dict] | None = None,
) -> bytes:
    liquidacion_items: list[tuple[ConciliacionItem, dict]] = []
    non_liquidacion_items: list[ConciliacionItem] = []
    for item in items:
        liq_meta = _extract_liquidacion_metadata(item)
        if liq_meta:
            liquidacion_items.append((item, liq_meta))
        else:
            non_liquidacion_items.append(item)

    has_liquidacion = bool(liquidacion_items)

    liquidacion_placas = {
        str(item.placa or "").strip().upper()
        for item, _ in liquidacion_items
        if str(item.placa or "").strip()
    }

    def _item_service_code_upper(item: ConciliacionItem) -> str:
        codigo_viaje = _item_servicio_codigo(item)
        if codigo_viaje:
            return codigo_viaje
        return str(item.servicio_codigo or "").strip().upper()

    def _item_service_label(item: ConciliacionItem) -> str:
        return (item.servicio_nombre or "").strip() or _item_service_code_upper(item) or str(getattr(item.tipo, "value", item.tipo))

    quincena_items: list[ConciliacionItem] = []
    additional_items: list[ConciliacionItem] = []
    for item in non_liquidacion_items:
        placa = str(item.placa or "").strip().upper()
        service_code = _item_service_code_upper(item)
        if service_code == "DISPONIBILIDAD":
            continue
        if has_liquidacion:
            if item.tipo == ItemTipo.VIAJE and service_code == "VIAJE" and placa in liquidacion_placas:
                quincena_items.append(item)
            else:
                additional_items.append(item)
        else:
            if service_code == "VIAJE":
                quincena_items.append(item)
            else:
                additional_items.append(item)

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_quincena = wb.create_sheet("Quincena")
    ws_adicionales = wb.create_sheet("Adicionales")

    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    header_font = Font(bold=True, color="1F2937")
    section_fill = PatternFill(fill_type="solid", fgColor="FDE68A")
    section_font = Font(bold=True, color="111827")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cop_format = '"$" #,##0'
    pct_format = '#,##0.##" %"'

    show_tarifa_tercero = user_role != UserRole.CLIENTE
    show_tarifa_cliente = user_role != UserRole.TERCERO
    show_cointra_financials = user_role == UserRole.COINTRA

    estado_display = _display_estado(conc)
    estado_label = estado_display.replace("_", " ")
    estado_styles = {
        "BORRADOR": ("ECFDF5", "065F46"),
        "EN_REVISION": ("FFFBEB", "92400E"),
        "APROBADA": ("F0FDFA", "115E59"),
        "ENVIADA_A_FACTURAR": ("F0F9FF", "075985"),
    }
    estado_fill_color, estado_font_color = estado_styles.get(estado_display, ("F7FEE7", "3F6212"))

    def _write_report_header(ws, title: str) -> int:
        row_num = 1
        ws.cell(row=row_num, column=1, value=f"Conciliacion #{conc.id} - {conc.nombre}")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12, color="111827")
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
        title_cell = ws.cell(row=row_num, column=5, value=title)
        title_cell.font = Font(bold=True, size=12, color="111827")
        estado_cell = ws.cell(row=row_num, column=7, value=f"ESTADO: {estado_label}")
        estado_cell.fill = PatternFill(fill_type="solid", fgColor=estado_fill_color)
        estado_cell.font = Font(bold=True, color=estado_font_color)
        estado_cell.alignment = center
        estado_cell.border = border
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}")
        ws.cell(row=row_num, column=1).font = Font(size=10, color="374151")
        return row_num + 2

    def _write_headers(ws, row_num: int, headers: list[str]) -> dict[str, int]:
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=idx, value=header)
            cell.fill = header_fill
            if header in {"Valor Cliente", "Valor Tercero", "Rentabilidad", "Ganancia Cointra"}:
                cell.fill = section_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        return {header: idx for idx, header in enumerate(headers, start=1)}

    def _style_row(ws, row_num: int, columns_count: int, fill: PatternFill | None = None, font: Font | None = None) -> None:
        for col_idx in range(1, columns_count + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = border
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font

    def _write_financials(
        ws,
        row_num: int,
        col_idx: dict[str, int],
        tarifa_cliente: float,
        tarifa_tercero: float,
    ) -> None:
        ganancia = tarifa_cliente - tarifa_tercero
        rentabilidad = (ganancia / tarifa_cliente * 100) if tarifa_cliente > 0 else 0.0
        if "Valor Cliente" in col_idx:
            c = ws.cell(row=row_num, column=col_idx["Valor Cliente"], value=tarifa_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in col_idx:
            c = ws.cell(row=row_num, column=col_idx["Valor Tercero"], value=tarifa_tercero)
            c.number_format = cop_format
        if "Rentabilidad" in col_idx:
            c = ws.cell(row=row_num, column=col_idx["Rentabilidad"], value=rentabilidad)
            c.number_format = pct_format
        if "Ganancia Cointra" in col_idx:
            c = ws.cell(row=row_num, column=col_idx["Ganancia Cointra"], value=ganancia)
            c.number_format = cop_format

    def _accumulate_totals(source_items: list[ConciliacionItem]) -> dict[str, dict[str, float]]:
        totals: dict[str, dict[str, float]] = {}
        for source_item in source_items:
            placa = str(source_item.placa or "").strip().upper() or "SIN_PLACA"
            bucket = totals.setdefault(placa, {"cliente": 0.0, "tercero": 0.0})
            bucket["cliente"] += _as_float(source_item.tarifa_cliente)
            bucket["tercero"] += _as_float(source_item.tarifa_tercero)
        return totals

    def _manifest_context(item: ConciliacionItem) -> dict[str, object]:
        manifiesto = _normalize_manifiesto_for_lookup(item.manifiesto_numero)
        avansat = (avansat_prefetched or {}).get(manifiesto) or {}
        has_manifiesto = bool(manifiesto)
        remesas = avansat.get("remesas") if isinstance(avansat.get("remesas"), list) else []
        remesas_rows = [row for row in remesas if isinstance(row, dict)]
        if not remesas_rows:
            remesas_rows = [
                {
                    "remesa": item.remesa or avansat.get("remesa") or "",
                    "producto": avansat.get("producto") or "",
                }
            ]
        return {
            "manifiesto": manifiesto,
            "has_manifiesto": has_manifiesto,
            "fecha_emision": avansat.get("fecha_emision") or str(item.fecha_servicio or ""),
            "placa": str(item.placa or "").strip().upper() or "SIN_PLACA",
            "trayler": avansat.get("trayler") or "",
            "ciudad_origen": avansat.get("ciudad_origen") or item.origen or "",
            "ciudad_destino": avansat.get("ciudad_destino") or item.destino or "",
            "remesas": remesas_rows,
        }

    def _write_adicionales_unified_section(
        ws,
        start_row: int,
        title: str,
        source_items: list[ConciliacionItem],
    ) -> tuple[int, dict[str, int], float, float, float, float]:
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11, color="111827")
        headers = [
            "Manifiesto",
            "Fecha Emision",
            "Placa Vehiculo",
            "Trayler",
            "Remesa",
            "Producto",
            "Ciudad Origen",
            "Ciudad Destino",
        ]
        if show_tarifa_cliente:
            headers.append("Valor Cliente")
        if show_tarifa_tercero:
            headers.append("Valor Tercero")
        if show_cointra_financials:
            headers.extend(["Rentabilidad", "Ganancia Cointra"])

        col_idx = _write_headers(ws, start_row + 1, headers)
        row_num = start_row + 2
        total_viajes_cliente = 0.0
        total_viajes_tercero = 0.0
        total_servicios_cliente = 0.0
        total_servicios_tercero = 0.0

        sorted_items = sorted(
            source_items,
            key=lambda current: (
                str(current.placa or "").strip().upper(),
                current.fecha_servicio,
                current.id,
            ),
        )

        if not sorted_items:
            ws.cell(row=row_num, column=1, value="(sin adicionales)")
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="6B7280")
            return row_num + 1, col_idx, total_viajes_cliente, total_viajes_tercero, total_servicios_cliente, total_servicios_tercero

        grouped_items: dict[str, list[ConciliacionItem]] = {}
        for item in sorted_items:
            placa = str(item.placa or "").strip().upper() or "SIN_PLACA"
            grouped_items.setdefault(placa, []).append(item)

        ordered_placas = sorted(grouped_items.keys())
        for placa_idx, placa in enumerate(ordered_placas):
            if placa_idx > 0:
                _write_headers(ws, row_num, headers)
                row_num += 1
            placa_total_cliente = 0.0
            placa_total_tercero = 0.0

            for item in grouped_items[placa]:
                manifest_data = _manifest_context(item)
                has_manifiesto = bool(manifest_data["has_manifiesto"])
                manifiesto_str = str(manifest_data["manifiesto"] or "")
                service_label = _item_service_label(item) if not has_manifiesto else ""

                tarifa_cliente = _as_float(item.tarifa_cliente)
                tarifa_tercero = _as_float(item.tarifa_tercero)
                placa_total_cliente += tarifa_cliente
                placa_total_tercero += tarifa_tercero

                if has_manifiesto:
                    total_viajes_cliente += tarifa_cliente
                    total_viajes_tercero += tarifa_tercero
                else:
                    total_servicios_cliente += tarifa_cliente
                    total_servicios_tercero += tarifa_tercero

                first_row = True
                for remesa_row in manifest_data["remesas"]:
                    values = {
                        "Manifiesto": manifiesto_str if first_row else "",
                        "Fecha Emision": manifest_data["fecha_emision"],
                        "Placa Vehiculo": placa,
                        "Trayler": manifest_data["trayler"],
                        "Remesa": str((remesa_row or {}).get("remesa") or "").strip(),
                        "Producto": service_label if (first_row and not has_manifiesto) else str((remesa_row or {}).get("producto") or "").strip(),
                        "Ciudad Origen": manifest_data["ciudad_origen"],
                        "Ciudad Destino": manifest_data["ciudad_destino"],
                    }
                    for header, column in col_idx.items():
                        if header in values:
                            ws.cell(row=row_num, column=column, value=values[header])
                    if first_row:
                        _write_financials(ws, row_num, col_idx, tarifa_cliente, tarifa_tercero)
                    _style_row(ws, row_num, len(headers))
                    row_num += 1
                    first_row = False

            ws.cell(row=row_num, column=col_idx["Ciudad Destino"], value=f"TOTAL PLACA {placa}")
            _write_financials(ws, row_num, col_idx, placa_total_cliente, placa_total_tercero)
            _style_row(ws, row_num, len(headers), fill=section_fill, font=section_font)
            row_num += 2

        return row_num, col_idx, total_viajes_cliente, total_viajes_tercero, total_servicios_cliente, total_servicios_tercero

    def _write_transport_section(
        ws,
        start_row: int,
        title: str,
        source_items: list[ConciliacionItem],
        write_plate_totals: bool = False,
        totals_label_prefix: str = "TOTAL PLACA",
    ) -> tuple[int, dict[str, int], float, float, dict[str, dict[str, float]]]:
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11, color="111827")
        headers = [
            "Manifiesto",
            "Fecha Emision",
            "Placa Vehiculo",
            "Trayler",
            "Remesa",
            "Producto",
            "Ciudad Origen",
            "Ciudad Destino",
        ]
        if show_tarifa_cliente:
            headers.append("Valor Cliente")
        if show_tarifa_tercero:
            headers.append("Valor Tercero")
        if show_cointra_financials:
            headers.extend(["Rentabilidad", "Ganancia Cointra"])

        col_idx = _write_headers(ws, start_row + 1, headers)
        row_num = start_row + 2
        total_cliente = 0.0
        total_tercero = 0.0
        totals_by_placa = _accumulate_totals(source_items)
        sorted_items = sorted(
            source_items,
            key=lambda current: (
                str(current.placa or "").strip().upper(),
                current.fecha_servicio,
                current.id,
            ),
        )

        if not sorted_items:
            ws.cell(row=row_num, column=1, value="(sin registros)")
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="6B7280")
            return row_num + 1, col_idx, total_cliente, total_tercero, totals_by_placa

        if write_plate_totals:
            grouped_items: dict[str, list[ConciliacionItem]] = {}
            for item in sorted_items:
                placa = str(item.placa or "").strip().upper() or "SIN_PLACA"
                grouped_items.setdefault(placa, []).append(item)

            ordered_placas = sorted(grouped_items.keys())
            for placa_idx, placa in enumerate(ordered_placas):
                if placa_idx > 0:
                    _write_headers(ws, row_num, headers)
                    row_num += 1
                for item in grouped_items[placa]:
                    manifest_data = _manifest_context(item)
                    tarifa_cliente = _as_float(item.tarifa_cliente)
                    tarifa_tercero = _as_float(item.tarifa_tercero)
                    total_cliente += tarifa_cliente
                    total_tercero += tarifa_tercero

                    first_row = True
                    for remesa_row in manifest_data["remesas"]:
                        values = {
                            "Manifiesto": manifest_data["manifiesto"],
                            "Fecha Emision": manifest_data["fecha_emision"],
                            "Placa Vehiculo": placa,
                            "Trayler": manifest_data["trayler"],
                            "Remesa": str((remesa_row or {}).get("remesa") or "").strip(),
                            "Producto": str((remesa_row or {}).get("producto") or "").strip(),
                            "Ciudad Origen": manifest_data["ciudad_origen"],
                            "Ciudad Destino": manifest_data["ciudad_destino"],
                        }
                        for header, column in col_idx.items():
                            if header in values:
                                ws.cell(row=row_num, column=column, value=values[header])
                        if first_row:
                            _write_financials(ws, row_num, col_idx, tarifa_cliente, tarifa_tercero)
                        _style_row(ws, row_num, len(headers))
                        row_num += 1
                        first_row = False

                placa_totals = totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
                ws.cell(row=row_num, column=col_idx["Ciudad Destino"], value=f"{totals_label_prefix} {placa}")
                _write_financials(
                    ws,
                    row_num,
                    col_idx,
                    placa_totals["cliente"],
                    placa_totals["tercero"],
                )
                _style_row(ws, row_num, len(headers), fill=section_fill, font=section_font)
                row_num += 2

            return row_num, col_idx, total_cliente, total_tercero, totals_by_placa

        current_placa = None
        for item in sorted_items:
            manifest_data = _manifest_context(item)
            placa = str(manifest_data["placa"])
            if current_placa is not None and placa != current_placa:
                row_num += 1  # fila en blanco de separación
                _write_headers(ws, row_num, headers)
                row_num += 1
            current_placa = placa

            tarifa_cliente = _as_float(item.tarifa_cliente)
            tarifa_tercero = _as_float(item.tarifa_tercero)
            total_cliente += tarifa_cliente
            total_tercero += tarifa_tercero

            first_row = True
            for remesa_row in manifest_data["remesas"]:
                values = {
                    "Manifiesto": manifest_data["manifiesto"],
                    "Fecha Emision": manifest_data["fecha_emision"],
                    "Placa Vehiculo": placa,
                    "Trayler": manifest_data["trayler"],
                    "Remesa": str((remesa_row or {}).get("remesa") or "").strip(),
                    "Producto": str((remesa_row or {}).get("producto") or "").strip(),
                    "Ciudad Origen": manifest_data["ciudad_origen"],
                    "Ciudad Destino": manifest_data["ciudad_destino"],
                }
                for header, column in col_idx.items():
                    if header in values:
                        ws.cell(row=row_num, column=column, value=values[header])
                if first_row:
                    _write_financials(ws, row_num, col_idx, tarifa_cliente, tarifa_tercero)
                _style_row(ws, row_num, len(headers))
                row_num += 1
                first_row = False

        return row_num, col_idx, total_cliente, total_tercero, totals_by_placa

    def _write_additional_services_section(
        ws,
        start_row: int,
        title: str,
        source_items: list[ConciliacionItem],
        write_plate_totals: bool = False,
        totals_label_prefix: str = "TOTAL PLACA",
    ) -> tuple[int, dict[str, int], float, float]:
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11, color="111827")
        headers = ["Placa", "Tipo Vehiculo", "Fecha", "Titulo Servicio", "Tipo Servicio"]
        if show_tarifa_cliente:
            headers.append("Valor Cliente")
        if show_tarifa_tercero:
            headers.append("Valor Tercero")
        if show_cointra_financials:
            headers.extend(["Rentabilidad", "Ganancia Cointra"])
        headers.append("Observaciones")

        col_idx = _write_headers(ws, start_row + 1, headers)
        row_num = start_row + 2
        total_cliente = 0.0
        total_tercero = 0.0
        sorted_items = sorted(
            source_items,
            key=lambda current: (
                str(current.placa or "").strip().upper(),
                current.fecha_servicio,
                current.id,
            ),
        )

        if not sorted_items:
            ws.cell(row=row_num, column=1, value="(sin servicios)")
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="6B7280")
            return row_num + 1, col_idx, total_cliente, total_tercero

        if write_plate_totals:
            grouped_items: dict[str, list[ConciliacionItem]] = {}
            for item in sorted_items:
                placa = str(item.placa or "").strip().upper() or "SIN_PLACA"
                grouped_items.setdefault(placa, []).append(item)

            ordered_placas = sorted(grouped_items.keys())
            for placa in ordered_placas:
                placa_total_cliente = 0.0
                placa_total_tercero = 0.0
                for item in grouped_items[placa]:
                    tarifa_cliente = _as_float(item.tarifa_cliente)
                    tarifa_tercero = _as_float(item.tarifa_tercero)
                    total_cliente += tarifa_cliente
                    total_tercero += tarifa_tercero
                    placa_total_cliente += tarifa_cliente
                    placa_total_tercero += tarifa_tercero
                    values = {
                        "Placa": placa,
                        "Tipo Vehiculo": tipo_vehiculo_by_placa.get(placa, ""),
                        "Fecha": str(item.fecha_servicio or ""),
                        "Titulo Servicio": item.viaje.titulo if item.viaje else "",
                        "Tipo Servicio": (item.servicio_nombre or "").strip()
                        or (item.servicio_codigo or "").strip()
                        or str(getattr(item.tipo, "value", item.tipo)),
                        "Observaciones": item.descripcion or "",
                    }
                    for header, column in col_idx.items():
                        if header in values:
                            ws.cell(row=row_num, column=column, value=values[header])
                    _write_financials(ws, row_num, col_idx, tarifa_cliente, tarifa_tercero)
                    _style_row(ws, row_num, len(headers))
                    row_num += 1

                ws.cell(row=row_num, column=col_idx["Tipo Vehiculo"], value=f"{totals_label_prefix} {placa}")
                _write_financials(ws, row_num, col_idx, placa_total_cliente, placa_total_tercero)
                _style_row(ws, row_num, len(headers), fill=section_fill, font=section_font)
                row_num += 2

            return row_num, col_idx, total_cliente, total_tercero

        for item in sorted_items:
            placa = str(item.placa or "").strip().upper() or "SIN_PLACA"
            tarifa_cliente = _as_float(item.tarifa_cliente)
            tarifa_tercero = _as_float(item.tarifa_tercero)
            total_cliente += tarifa_cliente
            total_tercero += tarifa_tercero
            values = {
                "Placa": placa,
                "Tipo Vehiculo": tipo_vehiculo_by_placa.get(placa, ""),
                "Fecha": str(item.fecha_servicio or ""),
                "Titulo Servicio": item.viaje.titulo if item.viaje else "",
                "Tipo Servicio": (item.servicio_nombre or "").strip()
                or (item.servicio_codigo or "").strip()
                or str(getattr(item.tipo, "value", item.tipo)),
                "Observaciones": item.descripcion or "",
            }
            for header, column in col_idx.items():
                if header in values:
                    ws.cell(row=row_num, column=column, value=values[header])
            _write_financials(ws, row_num, col_idx, tarifa_cliente, tarifa_tercero)
            _style_row(ws, row_num, len(headers))
            row_num += 1

        return row_num, col_idx, total_cliente, total_tercero

    liquidacion_items_sorted = sorted(
        liquidacion_items,
        key=lambda pair: (
            str(pair[0].placa or "").strip().upper(),
            pair[0].fecha_servicio,
            pair[0].id,
        ),
    )
    liquidacion_totals_by_placa = _accumulate_totals([item for item, _ in liquidacion_items_sorted])

    if has_liquidacion:
        current_row = _write_report_header(ws_resumen, "Resumen")
        ws_resumen.cell(row=current_row, column=1, value="LIQUIDACION CONTRATO FIJO")
        ws_resumen.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="111827")
        top_headers = ["Placa", "Tipo Vehiculo"]
        if show_tarifa_cliente:
            top_headers.append("Valor Cliente")
        if show_tarifa_tercero:
            top_headers.append("Valor Tercero")
        if show_cointra_financials:
            top_headers.extend(["Rentabilidad", "Ganancia Cointra"])
        top_col_idx = _write_headers(ws_resumen, current_row + 1, top_headers)
        current_row += 2

        top_total_cliente = 0.0
        top_total_tercero = 0.0
        for item, _ in liquidacion_items_sorted:
            placa = str(item.placa or "").strip().upper()
            tarifa_cliente = _as_float(item.tarifa_cliente)
            tarifa_tercero = _as_float(item.tarifa_tercero)
            top_total_cliente += tarifa_cliente
            top_total_tercero += tarifa_tercero
            ws_resumen.cell(row=current_row, column=top_col_idx["Placa"], value=placa)
            ws_resumen.cell(
                row=current_row,
                column=top_col_idx["Tipo Vehiculo"],
                value=tipo_vehiculo_by_placa.get(placa, ""),
            )
            _write_financials(ws_resumen, current_row, top_col_idx, tarifa_cliente, tarifa_tercero)
            _style_row(ws_resumen, current_row, len(top_headers))
            current_row += 1

        ws_resumen.cell(row=current_row, column=2, value="TOTAL CONTRATO FIJO")
        _write_financials(ws_resumen, current_row, top_col_idx, top_total_cliente, top_total_tercero)
        _style_row(ws_resumen, current_row, len(top_headers), fill=section_fill, font=section_font)
        current_row += 2

        summary_row, summary_bottom_idx, summary_total_cliente, summary_total_tercero = _write_additional_services_section(
            ws_resumen,
            current_row,
            "ADICIONALES",
            additional_items,
        )
        if additional_items:
            ws_resumen.cell(row=summary_row, column=2, value="TOTAL ADICIONALES")
            _write_financials(ws_resumen, summary_row, summary_bottom_idx, summary_total_cliente, summary_total_tercero)
            _style_row(
                ws_resumen,
                summary_row,
                len(summary_bottom_idx),
                fill=section_fill,
                font=section_font,
            )

        current_row = _write_report_header(ws_quincena, "Quincena")
        current_row, quincena_col_idx, _, _, quincena_totals_by_placa = _write_transport_section(
            ws_quincena,
            current_row,
            "VIAJES",
            quincena_items,
        )
        current_row += 2

        ws_quincena.cell(row=current_row, column=1, value="DISPONIBILIDAD")
        ws_quincena.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="111827")
        disponibilidad_headers = ["Placa", "Tipo Vehiculo"]
        if show_tarifa_cliente:
            disponibilidad_headers.append("Valor Cliente")
        if show_tarifa_tercero:
            disponibilidad_headers.append("Valor Tercero")
        if show_cointra_financials:
            disponibilidad_headers.append("Ganancia Cointra")
        disponibilidad_col_idx = _write_headers(ws_quincena, current_row + 1, disponibilidad_headers)
        current_row += 2

        for placa in sorted(liquidacion_totals_by_placa.keys()):
            liq_totals = liquidacion_totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
            viajes_totals = quincena_totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
            disponibilidad_cliente = liq_totals["cliente"] - viajes_totals["cliente"]
            disponibilidad_tercero = liq_totals["tercero"] - viajes_totals["tercero"]
            ws_quincena.cell(row=current_row, column=disponibilidad_col_idx["Placa"], value=placa)
            ws_quincena.cell(
                row=current_row,
                column=disponibilidad_col_idx["Tipo Vehiculo"],
                value=tipo_vehiculo_by_placa.get(placa, ""),
            )
            if "Valor Cliente" in disponibilidad_col_idx:
                c = ws_quincena.cell(
                    row=current_row,
                    column=disponibilidad_col_idx["Valor Cliente"],
                    value=disponibilidad_cliente,
                )
                c.number_format = cop_format
            if "Valor Tercero" in disponibilidad_col_idx:
                c = ws_quincena.cell(
                    row=current_row,
                    column=disponibilidad_col_idx["Valor Tercero"],
                    value=disponibilidad_tercero,
                )
                c.number_format = cop_format
            if "Ganancia Cointra" in disponibilidad_col_idx:
                c = ws_quincena.cell(
                    row=current_row,
                    column=disponibilidad_col_idx["Ganancia Cointra"],
                    value=disponibilidad_cliente - disponibilidad_tercero,
                )
                c.number_format = cop_format
            _style_row(ws_quincena, current_row, len(disponibilidad_headers))
            current_row += 1

        current_row += 2
        ws_quincena.cell(row=current_row, column=1, value="TOTALES POR VEHICULO")
        ws_quincena.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="111827")
        resumen_quincena_headers = ["Placa", "Concepto"]
        if show_tarifa_cliente:
            resumen_quincena_headers.append("Valor Cliente")
        if show_tarifa_tercero:
            resumen_quincena_headers.append("Valor Tercero")
        if show_cointra_financials:
            resumen_quincena_headers.append("Ganancia Cointra")
        resumen_quincena_col_idx = _write_headers(ws_quincena, current_row + 1, resumen_quincena_headers)
        current_row += 2

        for placa in sorted(liquidacion_totals_by_placa.keys()):
            liq_totals = liquidacion_totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
            viajes_totals = quincena_totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
            disponibilidad_cliente = liq_totals["cliente"] - viajes_totals["cliente"]
            disponibilidad_tercero = liq_totals["tercero"] - viajes_totals["tercero"]
            rows = [
                ("TOTAL VIAJES", viajes_totals["cliente"], viajes_totals["tercero"]),
                ("TOTAL DISPONIBILIDAD", disponibilidad_cliente, disponibilidad_tercero),
                ("TOTAL LIQUIDADO", viajes_totals["cliente"] + disponibilidad_cliente, viajes_totals["tercero"] + disponibilidad_tercero),
            ]
            for label, total_cliente, total_tercero in rows:
                ws_quincena.cell(row=current_row, column=resumen_quincena_col_idx["Placa"], value=placa)
                ws_quincena.cell(row=current_row, column=resumen_quincena_col_idx["Concepto"], value=label)
                if "Valor Cliente" in resumen_quincena_col_idx:
                    c = ws_quincena.cell(
                        row=current_row,
                        column=resumen_quincena_col_idx["Valor Cliente"],
                        value=total_cliente,
                    )
                    c.number_format = cop_format
                if "Valor Tercero" in resumen_quincena_col_idx:
                    c = ws_quincena.cell(
                        row=current_row,
                        column=resumen_quincena_col_idx["Valor Tercero"],
                        value=total_tercero,
                    )
                    c.number_format = cop_format
                if "Ganancia Cointra" in resumen_quincena_col_idx:
                    c = ws_quincena.cell(
                        row=current_row,
                        column=resumen_quincena_col_idx["Ganancia Cointra"],
                        value=total_cliente - total_tercero,
                    )
                    c.number_format = cop_format
                _style_row(ws_quincena, current_row, len(resumen_quincena_headers), font=section_font)
                current_row += 1
            current_row += 1

        current_row += 1
        ws_quincena.cell(row=current_row, column=1, value="CONSOLIDADO FACTURA")
        ws_quincena.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="111827")
        consolidated_headers = ["Concepto"]
        if show_tarifa_cliente:
            consolidated_headers.append("Valor Cliente")
        if show_tarifa_tercero:
            consolidated_headers.append("Valor Tercero")
        if show_cointra_financials:
            consolidated_headers.append("Ganancia Cointra")
        consolidated_col_idx = _write_headers(ws_quincena, current_row + 1, consolidated_headers)
        current_row += 2

        total_viajes_cliente = sum(values.get("cliente", 0.0) for values in quincena_totals_by_placa.values())
        total_viajes_tercero = sum(values.get("tercero", 0.0) for values in quincena_totals_by_placa.values())
        total_disponibilidad_cliente = 0.0
        total_disponibilidad_tercero = 0.0
        for placa in liquidacion_totals_by_placa.keys():
            liq_totals = liquidacion_totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
            viajes_totals = quincena_totals_by_placa.get(placa, {"cliente": 0.0, "tercero": 0.0})
            total_disponibilidad_cliente += liq_totals["cliente"] - viajes_totals["cliente"]
            total_disponibilidad_tercero += liq_totals["tercero"] - viajes_totals["tercero"]

        consolidated_rows = [
            ("FACTURA VIAJES", total_viajes_cliente, total_viajes_tercero),
            ("FACTURA DISPONIBILIDAD", total_disponibilidad_cliente, total_disponibilidad_tercero),
        ]
        for label, total_cliente, total_tercero in consolidated_rows:
            ws_quincena.cell(row=current_row, column=consolidated_col_idx["Concepto"], value=label)
            if "Valor Cliente" in consolidated_col_idx:
                c = ws_quincena.cell(
                    row=current_row,
                    column=consolidated_col_idx["Valor Cliente"],
                    value=total_cliente,
                )
                c.number_format = cop_format
            if "Valor Tercero" in consolidated_col_idx:
                c = ws_quincena.cell(
                    row=current_row,
                    column=consolidated_col_idx["Valor Tercero"],
                    value=total_tercero,
                )
                c.number_format = cop_format
            if "Ganancia Cointra" in consolidated_col_idx:
                c = ws_quincena.cell(
                    row=current_row,
                    column=consolidated_col_idx["Ganancia Cointra"],
                    value=total_cliente - total_tercero,
                )
                c.number_format = cop_format
            _style_row(ws_quincena, current_row, len(consolidated_headers), fill=section_fill, font=section_font)
            current_row += 1

        current_row += 1
        factura_total_cliente = total_viajes_cliente + total_disponibilidad_cliente
        factura_total_tercero = total_viajes_tercero + total_disponibilidad_tercero
        ws_quincena.cell(row=current_row, column=consolidated_col_idx["Concepto"], value="TOTAL FACTURA")
        if "Valor Cliente" in consolidated_col_idx:
            c = ws_quincena.cell(
                row=current_row,
                column=consolidated_col_idx["Valor Cliente"],
                value=factura_total_cliente,
            )
            c.number_format = cop_format
        if "Valor Tercero" in consolidated_col_idx:
            c = ws_quincena.cell(
                row=current_row,
                column=consolidated_col_idx["Valor Tercero"],
                value=factura_total_tercero,
            )
            c.number_format = cop_format
        if "Ganancia Cointra" in consolidated_col_idx:
            c = ws_quincena.cell(
                row=current_row,
                column=consolidated_col_idx["Ganancia Cointra"],
                value=factura_total_cliente - factura_total_tercero,
            )
            c.number_format = cop_format
        _style_row(ws_quincena, current_row, len(consolidated_headers), fill=section_fill, font=section_font)
    else:
        current_row = _write_report_header(ws_resumen, "Resumen")
        summary_row, summary_viajes_idx, summary_viajes_cliente, summary_viajes_tercero = _write_additional_services_section(
            ws_resumen,
            current_row,
            "VIAJES",
            quincena_items,
            write_plate_totals=True,
            totals_label_prefix="TOTAL VEHICULO",
        )
        ws_resumen.cell(row=summary_row, column=2, value="TOTAL VIAJES")
        _write_financials(ws_resumen, summary_row, summary_viajes_idx, summary_viajes_cliente, summary_viajes_tercero)
        _style_row(
            ws_resumen,
            summary_row,
            len(summary_viajes_idx),
            fill=section_fill,
            font=section_font,
        )

        if additional_items:
            current_row = summary_row + 2
            summary_row, summary_bottom_idx, summary_total_cliente, summary_total_tercero = _write_additional_services_section(
                ws_resumen,
                current_row,
                "ADICIONALES",
                additional_items,
                write_plate_totals=True,
                totals_label_prefix="TOTAL VEHICULO",
            )
            ws_resumen.cell(row=summary_row, column=2, value="TOTAL ADICIONALES")
            _write_financials(ws_resumen, summary_row, summary_bottom_idx, summary_total_cliente, summary_total_tercero)
            _style_row(
                ws_resumen,
                summary_row,
                len(summary_bottom_idx),
                fill=section_fill,
                font=section_font,
            )

        current_row = _write_report_header(ws_quincena, "Quincena")
        current_row, quincena_col_idx, total_viajes_cliente, total_viajes_tercero, _ = _write_transport_section(
            ws_quincena,
            current_row,
            "VIAJES",
            quincena_items,
        )

        current_row += 2
        consolidated_headers = ["Concepto"]
        if show_tarifa_cliente:
            consolidated_headers.append("Valor Cliente")
        if show_tarifa_tercero:
            consolidated_headers.append("Valor Tercero")
        if show_cointra_financials:
            consolidated_headers.append("Ganancia Cointra")
        consolidated_col_idx = _write_headers(ws_quincena, current_row, consolidated_headers)
        current_row += 1
        ws_quincena.cell(row=current_row, column=consolidated_col_idx["Concepto"], value="TOTAL VIAJES")
        if "Valor Cliente" in consolidated_col_idx:
            c = ws_quincena.cell(row=current_row, column=consolidated_col_idx["Valor Cliente"], value=total_viajes_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in consolidated_col_idx:
            c = ws_quincena.cell(row=current_row, column=consolidated_col_idx["Valor Tercero"], value=total_viajes_tercero)
            c.number_format = cop_format
        if "Ganancia Cointra" in consolidated_col_idx:
            c = ws_quincena.cell(row=current_row, column=consolidated_col_idx["Ganancia Cointra"], value=total_viajes_cliente - total_viajes_tercero)
            c.number_format = cop_format
        _style_row(ws_quincena, current_row, len(consolidated_headers), fill=section_fill, font=section_font)

    if additional_items:
        current_row = _write_report_header(ws_adicionales, "Adicionales")
        current_row, _, additional_trip_total_cliente, additional_trip_total_tercero, additional_services_total_cliente, additional_services_total_tercero = _write_adicionales_unified_section(
            ws_adicionales,
            current_row,
            "ADICIONALES",
            additional_items,
        )
        current_row += 2

        totals_headers = ["Concepto"]
        if show_tarifa_cliente:
            totals_headers.append("Valor Cliente")
        if show_tarifa_tercero:
            totals_headers.append("Valor Tercero")
        if show_cointra_financials:
            totals_headers.append("Ganancia Cointra")
        totals_col_idx = _write_headers(ws_adicionales, current_row, totals_headers)
        current_row += 1
        totals_rows = [
            ("FACTURA VIAJES", additional_trip_total_cliente, additional_trip_total_tercero),
            ("FACTURA SERVICIOS", additional_services_total_cliente, additional_services_total_tercero),
        ]
        for label, total_cliente, total_tercero in totals_rows:
            ws_adicionales.cell(row=current_row, column=totals_col_idx["Concepto"], value=label)
            if "Valor Cliente" in totals_col_idx:
                c = ws_adicionales.cell(row=current_row, column=totals_col_idx["Valor Cliente"], value=total_cliente)
                c.number_format = cop_format
            if "Valor Tercero" in totals_col_idx:
                c = ws_adicionales.cell(row=current_row, column=totals_col_idx["Valor Tercero"], value=total_tercero)
                c.number_format = cop_format
            if "Ganancia Cointra" in totals_col_idx:
                c = ws_adicionales.cell(row=current_row, column=totals_col_idx["Ganancia Cointra"], value=total_cliente - total_tercero)
                c.number_format = cop_format
            _style_row(ws_adicionales, current_row, len(totals_headers), fill=section_fill, font=section_font)
            current_row += 1

        current_row += 1
        total_adic_cliente = additional_trip_total_cliente + additional_services_total_cliente
        total_adic_tercero = additional_trip_total_tercero + additional_services_total_tercero
        ws_adicionales.cell(row=current_row, column=totals_col_idx["Concepto"], value="TOTAL ADICIONALES")
        if "Valor Cliente" in totals_col_idx:
            c = ws_adicionales.cell(row=current_row, column=totals_col_idx["Valor Cliente"], value=total_adic_cliente)
            c.number_format = cop_format
        if "Valor Tercero" in totals_col_idx:
            c = ws_adicionales.cell(row=current_row, column=totals_col_idx["Valor Tercero"], value=total_adic_tercero)
            c.number_format = cop_format
        if "Ganancia Cointra" in totals_col_idx:
            c = ws_adicionales.cell(row=current_row, column=totals_col_idx["Ganancia Cointra"], value=total_adic_cliente - total_adic_tercero)
            c.number_format = cop_format
        _style_row(ws_adicionales, current_row, len(totals_headers), fill=section_fill, font=section_font)
    else:
        wb.remove(ws_adicionales)

    resumen_widths = {
        "Placa": 16,
        "Tipo Vehiculo": 22,
        "Fecha": 16,
        "Titulo Servicio": 28,
        "Tipo Servicio": 24,
        "Valor Cliente": 18,
        "Valor Tercero": 18,
        "Rentabilidad": 16,
        "Ganancia Cointra": 20,
        "Observaciones": 36,
    }
    quincena_widths = {
        "Manifiesto": 18,
        "Fecha Emision": 18,
        "Placa Vehiculo": 16,
        "Trayler": 16,
        "Remesa": 18,
        "Producto": 28,
        "Ciudad Origen": 22,
        "Ciudad Destino": 22,
        "Valor Cliente": 18,
        "Valor Tercero": 18,
        "Rentabilidad": 16,
        "Ganancia Cointra": 20,
        "Placa": 16,
        "Tipo Vehiculo": 22,
        "Concepto": 24,
    }
    adicionales_widths = {**quincena_widths, **resumen_widths}

    sheets_with_widths = [
        (ws_resumen, resumen_widths),
        (ws_quincena, quincena_widths),
    ]
    if "Adicionales" in wb.sheetnames:
        sheets_with_widths.append((wb["Adicionales"], adicionales_widths))

    for sheet, widths in sheets_with_widths:
        min_column_width = 14
        max_column = sheet.max_column
        for col_idx in range(1, max_column + 1):
            header_value = str(sheet.cell(row=1, column=col_idx).value or "")
            best_width = widths.get(header_value, 18)
            for row_idx in range(1, min(sheet.max_row, 12) + 1):
                header_candidate = str(sheet.cell(row=row_idx, column=col_idx).value or "")
                if header_candidate in widths:
                    best_width = widths[header_candidate]
                    break
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(best_width, min_column_width)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@router.post("", response_model=ConciliacionOut)
def create_conciliacion(
    payload: ConciliacionCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    # Crear conciliaciones: COINTRA_ADMIN, COINTRA_USER (rol COINTRA)
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo usuarios Cointra pueden crear conciliaciones")

    operacion = db.get(Operacion, payload.operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    conc = Conciliacion(
        operacion_id=payload.operacion_id,
        nombre=payload.nombre,
        fecha_inicio=payload.fecha_inicio,
        fecha_fin=payload.fecha_fin,
        activo=True,
        borrador_guardado=False,
        enviada_facturacion=False,
        factura_cliente_enviada=False,
        created_by=user.id,
    )
    db.add(conc)
    db.flush()

    # Cargar automaticamente todos los viajes PENDIENTES de la operacion
    viajes_pendientes = (
        db.query(Viaje)
        .filter(
            Viaje.operacion_id == payload.operacion_id,
            Viaje.conciliacion_id.is_(None),
            Viaje.fecha_servicio <= payload.fecha_fin,
        )
        .order_by(Viaje.fecha_servicio.asc(), Viaje.id.asc())
        .all()
    )

    for viaje in viajes_pendientes:
        tarifa_tercero, tarifa_cliente, rentabilidad = _default_viaje_item_financials(viaje)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            viaje_id=viaje.id,
            tipo=ItemTipo.VIAJE,
            fecha_servicio=viaje.fecha_servicio,
            origen=viaje.origen,
            destino=viaje.destino,
            placa=viaje.placa,
            conductor=viaje.conductor,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=tarifa_cliente,
            rentabilidad=rentabilidad,
            manifiesto_numero=viaje.manifiesto_numero,
            remesa=None,
            descripcion=viaje.descripcion,
            created_by=user.id,
            cargado_por=viaje.cargado_por,
        )
        estado_valor = getattr(conc.estado, "value", conc.estado)
        viaje.conciliado = _should_mark_conciliado(estado_valor)
        viaje.estado_conciliacion = str(estado_valor)
        viaje.conciliacion_id = conc.id
        db.add(item)
        log_change(
            db,
            usuario_id=user.id,
            conciliacion_id=conc.id,
            campo="viaje_adjuntado",
            valor_nuevo=f"viaje_id={viaje.id}",
        )

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="conciliacion_creada",
        valor_nuevo=f"{payload.nombre} ({payload.fecha_inicio} - {payload.fecha_fin})",
    )
    db.commit()
    db.refresh(conc)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    create_internal_notifications(
        db,
        recipients,
        titulo="Nueva conciliacion creada",
        mensaje=f"Se creo la conciliacion '{conc.nombre}' para la operacion '{operacion.nombre}'.",
        tipo="CONCILIACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc

@router.get("", response_model=list[ConciliacionOut])
def list_conciliaciones(db: Session = Depends(get_db), user: Usuario = Depends(get_current_user)):
    query = (
        db.query(Conciliacion)
        .join(Operacion, Operacion.id == Conciliacion.operacion_id)
        .options(
            selectinload(Conciliacion.operacion).selectinload(Operacion.cliente),
            selectinload(Conciliacion.operacion).selectinload(Operacion.tercero),
            selectinload(Conciliacion.creador),
        )
    )
    if user.rol == UserRole.CLIENTE:
        query = query.join(
            usuario_operaciones_asignadas,
            usuario_operaciones_asignadas.c.operacion_id == Operacion.id,
        ).filter(
            usuario_operaciones_asignadas.c.usuario_id == user.id,
            Conciliacion.estado != "BORRADOR",
        )
    if user.rol == UserRole.TERCERO and user.tercero_id:
        query = query.filter(Operacion.tercero_id == user.tercero_id, Conciliacion.estado != "BORRADOR")
    if not is_cointra_admin(user):
        query = query.filter(Conciliacion.activo.is_(True))
    concs = query.order_by(Conciliacion.id.desc()).all()
    totals_map = _build_conciliacion_totals_map(db, [c.id for c in concs])
    return [_enrich_conciliacion(db, c, user, totals_map) for c in concs]


@router.get("/historial-cerradas", response_model=list[ConciliacionOut])
def list_closed_history(
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    cliente_id: int | None = None,
    tercero_id: int | None = None,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    query = (
        db.query(Conciliacion)
        .join(Operacion, Operacion.id == Conciliacion.operacion_id)
        .options(
            selectinload(Conciliacion.operacion).selectinload(Operacion.cliente),
            selectinload(Conciliacion.operacion).selectinload(Operacion.tercero),
            selectinload(Conciliacion.creador),
        )
        .filter(Conciliacion.estado == "CERRADA")
    )
    if user.rol == UserRole.CLIENTE:
        query = query.join(
            usuario_operaciones_asignadas,
            usuario_operaciones_asignadas.c.operacion_id == Operacion.id,
        ).filter(usuario_operaciones_asignadas.c.usuario_id == user.id)
    if user.rol == UserRole.TERCERO and user.tercero_id:
        query = query.filter(Operacion.tercero_id == user.tercero_id)
    if not is_cointra_admin(user):
        query = query.filter(Conciliacion.activo.is_(True))
    if cliente_id:
        query = query.filter(Operacion.cliente_id == cliente_id)
    if tercero_id:
        query = query.filter(Operacion.tercero_id == tercero_id)
    if fecha_inicio:
        query = query.filter(Conciliacion.fecha_inicio >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Conciliacion.fecha_fin <= fecha_fin)
    concs = query.order_by(Conciliacion.id.desc()).all()
    totals_map = _build_conciliacion_totals_map(db, [c.id for c in concs])
    return [_enrich_conciliacion(db, c, user, totals_map) for c in concs]


@router.patch("/{conciliacion_id}", response_model=ConciliacionOut)
def update_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionUpdate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No se enviaron cambios")

    if "operacion_id" in data:
        operacion = db.get(Operacion, data["operacion_id"])
        if not operacion or not operacion.activa:
            raise HTTPException(status_code=404, detail="Operacion no encontrada")

    for field, value in data.items():
        setattr(conc, field, value)

    db.commit()
    db.refresh(conc)
    return conc


@router.delete("/{conciliacion_id}")
def deactivate_conciliacion(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    conc.activo = False
    db.commit()
    return {"ok": True}


@router.post("/{conciliacion_id}/guardar-borrador", response_model=ConciliacionOut)
def guardar_conciliacion_borrador(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede guardar borradores")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo conciliaciones en BORRADOR se pueden guardar")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    conc.borrador_guardado = True
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conciliacion_id,
        campo="guardar_borrador",
        valor_nuevo="ok",
    )
    db.commit()
    db.refresh(conc)
    return conc


@router.post("/{conciliacion_id}/reactivar")
def reactivate_conciliacion(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    conc.activo = True
    db.commit()
    return {"ok": True}



@router.patch("/{conciliacion_id}/estado", response_model=ConciliacionOut)
def update_estado_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionUpdateEstado,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede cambiar estado de conciliacion")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_estado = conc.estado
    conc.estado = payload.estado
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="estado_conciliacion",
        valor_anterior=old_estado,
        valor_nuevo=payload.estado,
    )
    db.commit()
    db.refresh(conc)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA, UserRole.CLIENTE, UserRole.TERCERO])
    create_internal_notifications(
        db,
        recipients,
        titulo="Cambio de estado de conciliacion",
        mensaje=f"La conciliacion '{conc.nombre}' cambio a estado {conc.estado}.",
        tipo="ESTADO",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/items", response_model=ConciliacionItemOut)
def create_item(
    payload: ConciliacionItemCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede crear items")
    conc = db.get(Conciliacion, payload.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    item = ConciliacionItem(
        conciliacion_id=payload.conciliacion_id,
        tipo=payload.tipo,
        fecha_servicio=payload.fecha_servicio,
        origen=payload.origen,
        destino=payload.destino,
        placa=payload.placa,
        conductor=payload.conductor,
        tarifa_tercero=payload.tarifa_tercero,
        tarifa_cliente=payload.tarifa_cliente,
        manifiesto_numero=payload.manifiesto_numero,
        remesa=payload.remesa,
        descripcion=payload.descripcion,
        created_by=user.id,
        cargado_por=user.rol.value,
    )

    if user.rol in [UserRole.TERCERO, UserRole.COINTRA]:
        if item.tarifa_tercero and not item.tarifa_cliente:
            apply_rentabilidad(item, operacion)

    db.add(item)
    _mark_borrador_dirty(conc)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="item_creado",
        valor_nuevo=f"tipo={item.tipo}; fecha={item.fecha_servicio}",
    )
    db.commit()
    db.refresh(item)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    create_internal_notifications(
        db,
        recipients,
        titulo="Decision de cliente sobre item",
        mensaje=f"El cliente marco el item #{item.id} como {item.estado} en la conciliacion '{conc.nombre}'.",
        tipo="APROBACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return item


@router.get("/{conciliacion_id}/items", response_model=list[ConciliacionItemOut])
def list_items(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    if _repair_missing_viaje_items(db, conc, user.id):
        db.commit()

    items = (
        db.query(ConciliacionItem)
        .options(selectinload(ConciliacionItem.viaje).selectinload(Viaje.servicio))
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .order_by(ConciliacionItem.id.desc())
        .all()
    )

    enriched_items: list[dict] = []
    for item in items:
        payload = ConciliacionItemOut.model_validate(item).model_dump()
        liquidacion_meta = _extract_liquidacion_metadata(item)
        if liquidacion_meta:
            payload.update(liquidacion_meta)
        enriched_items.append(sanitize_item_for_role(payload, user.rol))

    return enriched_items


@router.post("/{conciliacion_id}/liquidacion-contrato-fijo", response_model=list[ConciliacionItemOut])
def create_liquidacion_contrato_fijo(
    conciliacion_id: int,
    payload: LiquidacionContratoFijoCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede crear liquidaciones de contrato fijo")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo conciliaciones en BORRADOR permiten agregar liquidaciones")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if payload.periodo_inicio > payload.periodo_fin:
        raise HTTPException(status_code=400, detail="El periodo de inicio no puede ser mayor al periodo final")

    liquidacion_id = payload.liquidacion_id
    if liquidacion_id is None:
        liquidacion_id = _next_liquidacion_id(db, conc.id)
    elif not _liquidacion_exists(db, conc.id, liquidacion_id):
        raise HTTPException(
            status_code=400,
            detail="La liquidacion seleccionada no existe en esta conciliacion. Crea una nueva primero.",
        )

    placas_norm = sorted({placa.strip().upper() for placa in payload.placas if placa and placa.strip()})
    if not placas_norm:
        raise HTTPException(status_code=400, detail="Debes seleccionar al menos una placa")

    vehiculos = (
        db.query(Vehiculo)
        .filter(
            Vehiculo.placa.in_(placas_norm),
            Vehiculo.activo.is_(True),
            Vehiculo.tercero_id == operacion.tercero_id,
        )
        .all()
    )
    vehiculos_by_placa = {v.placa.upper(): v for v in vehiculos}
    faltantes = [placa for placa in placas_norm if placa not in vehiculos_by_placa]
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=(
                "Las siguientes placas no estan disponibles para el tercero de la operacion: "
                + ", ".join(faltantes)
            ),
        )

    # Impedir duplicados: una placa solo puede tener un registro de liquidacion por conciliacion.
    existing_liq_items = (
        db.query(ConciliacionItem)
        .filter(ConciliacionItem.conciliacion_id == conc.id)
        .all()
    )
    existing_liq_placas = {
        str(i.placa or "").strip().upper()
        for i in existing_liq_items
        if _extract_liquidacion_metadata(i)
    }
    duplicadas = [p for p in placas_norm if p in existing_liq_placas]
    if duplicadas:
        raise HTTPException(
            status_code=400,
            detail=(
                "Ya existe un registro de liquidacion contrato fijo para la(s) placa(s): "
                + ", ".join(duplicadas)
            ),
        )

    created_rows: list[ConciliacionItem] = []
    for placa in placas_norm:
        tarifa_tercero = float(payload.valor_tercero)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            tipo=ItemTipo.OTRO,
            fecha_servicio=payload.periodo_fin,
            origen="Liquidacion Contrato Fijo",
            destino="Liquidacion Contrato Fijo",
            placa=placa,
            conductor=None,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=None,
            rentabilidad=None,
            manifiesto_numero=None,
            remesa=None,
            descripcion=_build_liquidacion_metadata(
                liquidacion_id,
                payload.periodo_inicio,
                payload.periodo_fin,
            ),
            created_by=user.id,
            cargado_por=user.rol.value,
        )
        apply_rentabilidad(item, operacion)
        db.add(item)
        db.flush()
        created_rows.append(item)

        log_change(
            db,
            usuario_id=user.id,
            conciliacion_id=conc.id,
            item_id=item.id,
            campo="liquidacion_contrato_fijo_creada",
            valor_nuevo=(
                f"liquidacion_id={liquidacion_id}; placa={placa}; periodo={payload.periodo_inicio} a {payload.periodo_fin}; "
                f"valor_tercero={tarifa_tercero}"
            ),
        )

    _mark_borrador_dirty(conc)

    db.commit()
    for row in created_rows:
        db.refresh(row)

    result: list[dict] = []
    for row in created_rows:
        row_payload = ConciliacionItemOut.model_validate(row).model_dump()
        meta = _extract_liquidacion_metadata(row)
        if meta:
            row_payload.update(meta)
        result.append(sanitize_item_for_role(row_payload, user.rol))

    return result


@router.delete("/items/{item_id}")
def delete_liquidacion_item(
    item_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede eliminar registros de liquidacion")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    conc = db.get(Conciliacion, item.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo en BORRADOR se pueden eliminar registros de liquidacion")

    liq_meta = _extract_liquidacion_metadata(item)
    if not liq_meta:
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar registros del bloque contrato fijo")

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=None,
        campo="liquidacion_contrato_fijo_eliminada",
        valor_anterior=f"item_id={item.id}; placa={item.placa}; t3={item.tarifa_tercero}; tc={item.tarifa_cliente}",
    )
    _mark_borrador_dirty(conc)
    db.query(HistorialCambio).filter(HistorialCambio.item_id == item.id).update(
        {HistorialCambio.item_id: None},
        synchronize_session=False,
    )

    # Buscar y eliminar el item de Disponibilidad de la misma placa en esta conciliacion
    placa_norm = str(item.placa or "").strip().upper()
    if placa_norm:
        disp_items = (
            db.query(ConciliacionItem)
            .filter(
                ConciliacionItem.conciliacion_id == conc.id,
                ConciliacionItem.id != item.id,
            )
            .all()
        )
        for d_item in disp_items:
            if str(d_item.placa or "").strip().upper() != placa_norm:
                continue
            if not d_item.viaje_id:
                continue
            viaje = db.get(Viaje, d_item.viaje_id)
            if not viaje:
                continue
            servicio = db.get(Servicio, viaje.servicio_id) if viaje.servicio_id else None
            es_disp = (
                servicio and (
                    str(servicio.codigo or "").strip().upper() == "DISPONIBILIDAD"
                    or str(servicio.nombre or "").strip().lower() == "disponibilidad"
                )
            ) or "disponibilidad" in str(viaje.titulo or "").lower()
            if not es_disp:
                continue
            db.query(HistorialCambio).filter(HistorialCambio.item_id == d_item.id).update(
                {HistorialCambio.item_id: None}, synchronize_session=False
            )
            db.delete(d_item)
            viaje.conciliacion_id = None
            viaje.estado_conciliacion = None
            viaje.activo = False

    db.delete(item)
    db.commit()
    return {"ok": True}


@router.delete("/{conciliacion_id}/disponibilidad/{item_id}")
def delete_disponibilidad_item(
    conciliacion_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Elimina un item de servicio Disponibilidad (auto-creado) y desactiva el viaje subyacente."""
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede eliminar items de disponibilidad")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo en BORRADOR se pueden eliminar items de disponibilidad")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    item = db.get(ConciliacionItem, item_id)
    if not item or item.conciliacion_id != conciliacion_id:
        raise HTTPException(status_code=404, detail="Item no encontrado en esta conciliacion")

    if not item.viaje_id:
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar items de servicio Disponibilidad")

    viaje = db.get(Viaje, item.viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje subyacente no encontrado")

    servicio = db.get(Servicio, viaje.servicio_id) if viaje.servicio_id else None
    es_disponibilidad = (
        servicio and (
            str(servicio.codigo or "").strip().upper() == "DISPONIBILIDAD"
            or str(servicio.nombre or "").strip().lower() == "disponibilidad"
        )
    ) or "disponibilidad" in str(viaje.titulo or "").lower()

    if not es_disponibilidad:
        raise HTTPException(
            status_code=400,
            detail="Solo se pueden eliminar items de servicio Disponibilidad creados automaticamente",
        )

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conciliacion_id,
        campo="disponibilidad_eliminada",
        valor_anterior=f"item_id={item.id}; viaje_id={viaje.id}; placa={item.placa}; t3={item.tarifa_tercero}",
    )

    # Limpiar FK del historial antes de borrar el item
    db.query(HistorialCambio).filter(HistorialCambio.item_id == item.id).update(
        {HistorialCambio.item_id: None}, synchronize_session=False
    )
    db.delete(item)

    # Desactivar el viaje auto-creado
    viaje.conciliacion_id = None
    viaje.estado_conciliacion = None
    viaje.conciliado = False
    viaje.activo = False

    _mark_borrador_dirty(conc)
    db.commit()
    return {"ok": True}


@router.patch("/items/{item_id}/estado", response_model=ConciliacionItemOut)
def update_item_estado(
    item_id: int,
    payload: ConciliacionItemUpdateEstado,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede cambiar estado de items")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    conc = db.get(Conciliacion, item.conciliacion_id)
    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_estado = item.estado
    item.estado = payload.estado
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=item.id,
        campo="estado_item",
        valor_anterior=old_estado,
        valor_nuevo=payload.estado,
    )

    db.commit()
    db.refresh(item)
    return item


@router.patch("/items/{item_id}", response_model=ConciliacionItemOut)
def patch_item(
    item_id: int,
    payload: ConciliacionItemPatch,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede actualizar items")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    conc = db.get(Conciliacion, item.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    can_edit_borrador = conc.estado == "BORRADOR"
    can_fix_manifiesto_aprobada = conc.estado == "APROBADA" and not conc.enviada_facturacion
    if not can_edit_borrador and not can_fix_manifiesto_aprobada:
        raise HTTPException(status_code=400, detail="Solo se puede editar en BORRADOR")

    changed = payload.model_fields_set
    if can_fix_manifiesto_aprobada:
        if not _is_transport_item(item):
            raise HTTPException(
                status_code=400,
                detail="Solo servicios de transporte (VIAJE/VIAJE_ADICIONAL) permiten ajustar manifiesto en estado APROBADA",
            )
        if not changed or not changed.issubset({"manifiesto_numero"}):
            raise HTTPException(
                status_code=400,
                detail="En APROBADA solo puedes corregir el manifiesto para enviar a facturacion",
            )

    if can_edit_borrador and changed:
        _mark_borrador_dirty(conc)

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_manifiesto = item.manifiesto_numero
    old_remesa = item.remesa
    old_fecha_servicio = item.fecha_servicio
    old_origen = item.origen
    old_destino = item.destino
    old_placa = item.placa
    old_conductor = item.conductor
    old_tarifa_tercero = item.tarifa_tercero
    old_tarifa_cliente = item.tarifa_cliente
    old_rentabilidad = item.rentabilidad
    old_descripcion = item.descripcion

    if "fecha_servicio" in changed:
        if payload.fecha_servicio is None:
            raise HTTPException(status_code=400, detail="La fecha del servicio es obligatoria")
        item.fecha_servicio = payload.fecha_servicio

    if "origen" in changed:
        normalized_origen = str(payload.origen or "").strip()
        item.origen = normalized_origen or None

    if "destino" in changed:
        normalized_destino = str(payload.destino or "").strip()
        item.destino = normalized_destino or None

    if "placa" in changed:
        normalized_placa = str(payload.placa or "").strip().upper()
        item.placa = normalized_placa or None

    if "conductor" in changed:
        normalized_conductor = str(payload.conductor or "").strip()
        item.conductor = normalized_conductor or None

    if "manifiesto_numero" in changed:
        item.manifiesto_numero = payload.manifiesto_numero
    if "remesa" in changed:
        item.remesa = payload.remesa
    if "descripcion" in changed:
        normalized_descripcion = str(payload.descripcion or "").strip()
        item.descripcion = normalized_descripcion or None

    pct = float(operacion.porcentaje_rentabilidad)
    # Usar rentabilidad actual del ítem; solo como fallback la de la operación
    pct = float(item.rentabilidad) if item.rentabilidad is not None else float(operacion.porcentaje_rentabilidad)

    tarifa_fields = changed & {"tarifa_tercero", "tarifa_cliente", "rentabilidad"}
    if tarifa_fields:
        if "tarifa_tercero" in changed and "tarifa_cliente" not in changed and "rentabilidad" not in changed:
            # Modificó tarifa_tercero → recalcular tarifa_cliente; rentabilidad no cambia
            item.tarifa_tercero = payload.tarifa_tercero
            if pct < 100:
                item.tarifa_cliente = payload.tarifa_tercero / (1 - pct / 100)
        elif "tarifa_cliente" in changed and "tarifa_tercero" not in changed and "rentabilidad" not in changed:
            # Modificó tarifa_cliente → recalcular tarifa_tercero; rentabilidad no cambia
            item.tarifa_cliente = payload.tarifa_cliente
            item.tarifa_tercero = payload.tarifa_cliente * (1 - pct / 100)
        elif "rentabilidad" in changed:
            # Modificó % rentabilidad → guardar nuevo %, recalcular tarifa_tercero; tarifa_cliente no cambia
            new_pct = payload.rentabilidad if payload.rentabilidad is not None else pct
            item.rentabilidad = new_pct
            if item.tarifa_cliente is not None and new_pct < 100:
                item.tarifa_tercero = float(item.tarifa_cliente) * (1 - new_pct / 100)

    if {"manifiesto_numero", "placa"} & changed:
        # Solo validar si el manifiesto tiene valor (no está siendo eliminado)
        if item.manifiesto_numero:
            _validate_transport_item_manifest_or_raise(db, item)

    # Mantener viaje sincronizado con la corrección final de conciliación para evitar divergencias.
    if item.viaje_id is not None:
        viaje = db.get(Viaje, item.viaje_id)
        if viaje:
            if "fecha_servicio" in changed and item.fecha_servicio:
                viaje.fecha_servicio = item.fecha_servicio
            if "origen" in changed:
                viaje.origen = item.origen or ""
            if "destino" in changed:
                viaje.destino = item.destino or ""
            if "placa" in changed and item.placa:
                viaje.placa = item.placa
            if "conductor" in changed:
                viaje.conductor = item.conductor
            if "manifiesto_numero" in changed:
                viaje.manifiesto_numero = item.manifiesto_numero
            if "descripcion" in changed:
                viaje.descripcion = item.descripcion
            if tarifa_fields:
                if item.tarifa_tercero is not None:
                    viaje.tarifa_tercero = item.tarifa_tercero
                if item.tarifa_cliente is not None:
                    viaje.tarifa_cliente = item.tarifa_cliente
                if item.rentabilidad is not None:
                    viaje.rentabilidad = item.rentabilidad

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=item.id,
        campo="actualizacion_manual_item",
        valor_anterior=(
            f"fecha={old_fecha_servicio}; origen={old_origen}; destino={old_destino}; "
            f"placa={old_placa}; conductor={old_conductor}; manifiesto={old_manifiesto}; "
            f"remesa={old_remesa}; t3={old_tarifa_tercero}; tc={old_tarifa_cliente}; "
            f"rent={old_rentabilidad}; descripcion={old_descripcion}"
        ),
        valor_nuevo=(
            f"fecha={item.fecha_servicio}; origen={item.origen}; destino={item.destino}; "
            f"placa={item.placa}; conductor={item.conductor}; manifiesto={item.manifiesto_numero}; "
            f"remesa={item.remesa}; t3={item.tarifa_tercero}; tc={item.tarifa_cliente}; "
            f"rent={item.rentabilidad}; descripcion={item.descripcion}"
        ),
    )

    db.commit()
    db.refresh(item)
    item_payload = ConciliacionItemOut.model_validate(item).model_dump()
    liquidacion_meta = _extract_liquidacion_metadata(item)
    if liquidacion_meta:
        item_payload.update(liquidacion_meta)
    return sanitize_item_for_role(item_payload, user.rol)


@router.patch("/items/{item_id}/decision-cliente", response_model=ConciliacionItemOut)
def cliente_decide_item(
    item_id: int,
    payload: ClienteItemDecision,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.CLIENTE:
        raise HTTPException(status_code=403, detail="Solo Cliente puede aprobar/rechazar items")
    if payload.estado not in [ItemEstado.APROBADO, ItemEstado.RECHAZADO]:
        raise HTTPException(status_code=400, detail="Estado permitido para Cliente: APROBADO o RECHAZADO")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    conc = db.get(Conciliacion, item.conciliacion_id)
    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_estado = item.estado
    item.estado = payload.estado
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=item.id,
        campo="decision_cliente_item",
        valor_anterior=old_estado,
        valor_nuevo=payload.estado,
    )
    if payload.comentario:
        db.add(
            Comentario(
                conciliacion_id=conc.id,
                item_id=item.id,
                usuario_id=user.id,
                comentario=payload.comentario,
            )
        )

    db.commit()
    db.refresh(item)
    return item


@router.post("/{conciliacion_id}/enviar-revision", response_model=ConciliacionOut)
def enviar_revision(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede enviar a revision")
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if conc.estado == "BORRADOR" and not conc.borrador_guardado:
        raise HTTPException(
            status_code=400,
            detail="Debes guardar la conciliacion antes de enviarla a revision",
        )

    items = (
        db.query(ConciliacionItem)
        .options(selectinload(ConciliacionItem.viaje).selectinload(Viaje.servicio))
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .order_by(ConciliacionItem.id.asc())
        .all()
    )
    _validate_transport_items_manifests_or_raise(db, items, "enviar a revision")

    conc.estado = "EN_REVISION"
    conc.enviada_facturacion = False
    conc.factura_cliente_enviada = False
    conc.po_numero_autorizacion = None
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="enviar_revision",
        valor_nuevo=payload.observacion or "sin observacion",
    )

    recipients = _resolve_recipients(db, operacion, [UserRole.CLIENTE])
    target_emails = _parse_target_emails(payload.destinatario_email, recipients)
    notification_recipients = _users_matching_emails(recipients, target_emails) or recipients

    if not target_emails:
        raise HTTPException(status_code=400, detail="No hay correo destinatario para enviar la conciliacion")

    if target_emails:
        subject = conc.nombre
        custom_message = payload.mensaje or ""
        login_url = _login_url()
        body = (
            f"Hola,\n\n"
            f"Cointra envio la conciliacion '{conc.nombre}' para tu revision.\n"
            f"Operacion: {operacion.nombre}\n"
            f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n\n"
            f"Enviado por: {_sender_signature(user)}\n\n"
            f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
            "Ingresa al sistema para revisar y autorizar la conciliacion.\n"
            f"Accede aqui: {login_url}\n\n"
        )
        email_result = send_manual_email(target_emails, subject=subject, body=body)
        if email_result["failed"] >= len(target_emails):
            db.rollback()
            detail = "No se pudo enviar el correo de revision"
            if email_result["errors"]:
                detail = f"{detail}: {email_result['errors'][0]}"
            raise HTTPException(status_code=502, detail=detail)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Conciliacion enviada a revision",
        mensaje=f"Cointra envio la conciliacion '{conc.nombre}' para tu revision.",
        tipo="ESTADO",
        conciliacion_id=conc.id,
    )

    db.commit()
    db.refresh(conc)
    return conc


@router.post("/{conciliacion_id}/aprobar-cliente", response_model=ConciliacionOut)
def aprobar_conciliacion_cliente(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.CLIENTE:
        raise HTTPException(status_code=403, detail="Solo Cliente puede aprobar conciliacion")
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)
    items = db.query(ConciliacionItem).filter(ConciliacionItem.conciliacion_id == conc.id).all()
    pendientes = [i for i in items if i.estado != ItemEstado.APROBADO]
    if pendientes:
        raise HTTPException(status_code=400, detail="No se puede aprobar: existen items no aprobados")

    po_numero = (payload.po_numero or "").strip()
    if not po_numero:
        raise HTTPException(status_code=400, detail="Debes registrar el número de PO para aprobar la conciliación")

    conc.estado = "APROBADA"
    conc.enviada_facturacion = False
    conc.factura_cliente_enviada = False
    conc.po_numero_autorizacion = po_numero
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="aprobacion_cliente",
        valor_nuevo=(
            f"{payload.observacion or 'aprobada por cliente'}"
            + (f" | PO: {po_numero}" if po_numero else "")
        ),
    )
    db.commit()
    db.refresh(conc)

    cointra_recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    tercero_recipients = _resolve_recipients(db, operacion, [UserRole.TERCERO])
    last_review_sender = _find_last_review_sender(db, conc.id)
    preferred_cointra = [last_review_sender] if last_review_sender and last_review_sender.rol == UserRole.COINTRA else []
    email_recipients = preferred_cointra or cointra_recipients
    target_emails = _parse_target_emails(payload.destinatario_email, email_recipients)
    if not target_emails:
        raise HTTPException(status_code=400, detail="Debes indicar un destinatario de correo para confirmar la aprobación")
    notification_recipients = list(email_recipients)
    for tercero in tercero_recipients:
        if all(existing.id != tercero.id for existing in notification_recipients):
            notification_recipients.append(tercero)
    subject = f"Conciliacion aprobada: {conc.nombre}"
    custom_message = payload.mensaje or ""
    login_url = _login_url()
    body = (
        f"Hola,\n\n"
        f"El cliente aprobo la conciliacion '{conc.nombre}'.\n"
        f"Operacion: {operacion.nombre}\n"
        f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n\n"
        f"PO autorizacion: {po_numero}\n\n"
        f"Enviado por: {_sender_signature(user)}\n\n"
        f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
        "Ingresa al sistema para continuar con el flujo.\n"
        f"Accede aqui: {login_url}\n\n"
    )
    send_manual_email(target_emails, subject=subject, body=body)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Conciliacion aprobada por cliente",
        mensaje=f"La conciliacion '{conc.nombre}' fue aprobada por el cliente y quedo autorizada para facturar.",
        tipo="APROBACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/{conciliacion_id}/devolver-cliente", response_model=ConciliacionOut)
def devolver_conciliacion_cliente(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.CLIENTE:
        raise HTTPException(status_code=403, detail="Solo Cliente puede devolver conciliacion")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    items = db.query(ConciliacionItem).filter(ConciliacionItem.conciliacion_id == conc.id).all()
    rechazados = [i for i in items if i.estado == ItemEstado.RECHAZADO]
    if not rechazados:
        raise HTTPException(
            status_code=400,
            detail="Para devolver la conciliacion debes rechazar al menos un item",
        )

    if not payload.observacion or not payload.observacion.strip():
        raise HTTPException(status_code=400, detail="Debes incluir observaciones para devolver")

    conc.estado = "BORRADOR"
    conc.enviada_facturacion = False
    conc.factura_cliente_enviada = False
    conc.po_numero_autorizacion = None
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="devolucion_cliente",
        valor_nuevo=payload.observacion,
    )
    db.add(
        Comentario(
            conciliacion_id=conc.id,
            usuario_id=user.id,
            comentario=payload.observacion,
        )
    )
    db.commit()
    db.refresh(conc)

    cointra_recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    last_review_sender = _find_last_review_sender(db, conc.id)
    email_recipients = [last_review_sender] if last_review_sender and last_review_sender.rol == UserRole.COINTRA else cointra_recipients
    target_emails = _parse_target_emails(payload.destinatario_email, email_recipients)
    notification_recipients = list(email_recipients)
    if target_emails:
        subject = f"Conciliacion devuelta con novedades: {conc.nombre}"
        custom_message = payload.mensaje or ""
        login_url = _login_url()
        body = (
            f"Hola,\n\n"
            f"El cliente devolvio la conciliacion '{conc.nombre}' con novedades.\n"
            f"Operacion: {operacion.nombre}\n"
            f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n"
            f"Observacion: {payload.observacion}\n\n"
            f"Enviado por: {_sender_signature(user)}\n\n"
            f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
            "Ingresa al sistema para revisar, ajustar y reenviar.\n"
            f"Accede aqui: {login_url}\n\n"
        )
        send_manual_email(target_emails, subject=subject, body=body)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Conciliacion devuelta con novedades",
        mensaje=f"El cliente devolvio la conciliacion '{conc.nombre}' con observaciones para revisar.",
        tipo="DEVOLUCION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/{conciliacion_id}/enviar-facturacion", response_model=ConciliacionOut)
def enviar_facturacion_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede enviar a facturacion")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    if conc.estado != "APROBADA":
        raise HTTPException(status_code=400, detail="Solo conciliaciones aprobadas pueden enviarse a facturacion")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    items = (
        db.query(ConciliacionItem)
        .options(selectinload(ConciliacionItem.viaje).selectinload(Viaje.servicio))
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .order_by(ConciliacionItem.id.asc())
        .all()
    )
    if not items:
        raise HTTPException(status_code=400, detail="No hay registros para generar el archivo de facturacion")

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    target_emails = _parse_target_emails(payload.destinatario_email, recipients)
    if not target_emails:
        raise HTTPException(status_code=400, detail="No hay correos de destino para facturacion")

    _validate_transport_items_manifests_or_raise(db, items, "enviar a facturacion")

    avansat_prefetched = _prefetch_avansat_for_manifest_numbers_or_raise(
        db,
        [str(item.manifiesto_numero or "") for item in items],
    )

    placas = {str(item.placa or "").strip().upper() for item in items if item.placa}
    tipo_vehiculo_by_placa: dict[str, str] = {}
    if placas:
        vehiculos = (
            db.query(Vehiculo)
            .options(selectinload(Vehiculo.tipo))
            .filter(Vehiculo.placa.in_(list(placas)))
            .all()
        )
        tipo_vehiculo_by_placa = {
            str(v.placa or "").strip().upper(): (v.tipo.nombre if v.tipo else "")
            for v in vehiculos
        }

    excel_content = _build_conciliacion_excel(
        conc,
        items,
        user.rol,
        tipo_vehiculo_by_placa,
        avansat_prefetched,
    )
    filename = f"conciliacion_{conc.id}_resumen.xlsx"
    custom_message = payload.mensaje or ""
    po_numero = (conc.po_numero_autorizacion or "").strip()
    email_body = (
        f"Hola,\n\n"
        f"Se envio la conciliacion '{conc.nombre}' para facturacion.\n"
        f"Operacion: {operacion.nombre}\n"
        f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n\n"
        f"PO autorizacion cliente: {po_numero or '(sin PO registrada)'}\n\n"
        f"Enviado por: {_sender_signature(user)}\n\n"
        f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
        "Adjunto encontraras el archivo Excel con los viajes.\n"
    )

    send_result = send_manual_email(
        target_emails,
        subject=f"Autorizacion para facturar: {conc.nombre}",
        body=email_body,
        attachments=[
            {
                "filename": filename,
                "content": excel_content,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        ],
    )
    if send_result["failed"] >= len(target_emails):
        detail = "No se pudo enviar el correo de facturacion"
        if send_result["errors"]:
            detail = f"{detail}: {send_result['errors'][0]}"
        raise HTTPException(status_code=502, detail=detail)

    conc.enviada_facturacion = True
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="envio_facturacion",
        valor_nuevo=f"destinatarios={', '.join(target_emails)}",
    )
    db.commit()
    db.refresh(conc)

    create_internal_notifications(
        db,
        recipients,
        titulo="Conciliacion enviada a facturar",
        mensaje=f"La conciliacion '{conc.nombre}' fue enviada a facturacion con archivo Excel adjunto.",
        tipo="FACTURACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/{conciliacion_id}/enviar-factura-cliente", response_model=ConciliacionOut)
def enviar_factura_cliente_conciliacion(
    conciliacion_id: int,
    destinatario_email: str | None = Form(default=None),
    mensaje: str | None = Form(default=None),
    archivos_factura: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede enviar factura al cliente")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    if conc.estado != "APROBADA" or not conc.enviada_facturacion:
        raise HTTPException(
            status_code=400,
            detail="La conciliacion debe estar enviada a facturacion antes de enviar la factura al cliente",
        )
    if conc.factura_cliente_enviada:
        raise HTTPException(status_code=400, detail="La factura ya fue enviada al cliente")

    if not archivos_factura:
        raise HTTPException(status_code=400, detail="Debes adjuntar al menos un archivo PDF de la factura")

    archivos_leidos: list[dict] = []
    for archivo in archivos_factura:
        if not archivo.filename:
            raise HTTPException(status_code=400, detail="Uno de los archivos no tiene nombre")
        fname = archivo.filename.strip()
        if not fname.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"El archivo '{fname}' debe estar en formato PDF")
        content = archivo.file.read()
        if not content:
            raise HTTPException(status_code=400, detail=f"El archivo '{fname}' está vacío")
        archivos_leidos.append({"filename": fname, "content": content})

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    recipients = _resolve_recipients(db, operacion, [UserRole.CLIENTE])
    target_emails = _parse_target_emails(destinatario_email, recipients)
    notification_recipients = _users_matching_emails(recipients, target_emails) or recipients
    if not target_emails:
        raise HTTPException(status_code=400, detail="No hay correo destinatario para enviar la factura")

    custom_message = (mensaje or "").strip()
    login_url = _login_url()
    body = (
        f"Hola,\n\n"
        f"Compartimos la factura en PDF de la conciliacion '{conc.nombre}' (#{conc.id}).\n"
        f"Operacion: {operacion.nombre}\n"
        f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n"
        f"PO autorizacion: {conc.po_numero_autorizacion or '(sin PO reportada)'}\n\n"
        f"Enviado por: {_sender_signature(user)}\n\n"
        f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
        "Adjunto encontraras la factura en formato PDF.\n"
        f"Accede aqui: {login_url}\n\n"
    )

    send_result = send_manual_email(
        target_emails,
        subject=f"Factura conciliacion {conc.nombre} #{conc.id}",
        body=body,
        attachments=[
            {
                "filename": a["filename"],
                "content": a["content"],
                "mime_type": "application/pdf",
            }
            for a in archivos_leidos
        ],
    )
    if send_result["failed"] >= len(target_emails):
        detail = "No se pudo enviar el correo de factura al cliente"
        if send_result["errors"]:
            detail = f"{detail}: {send_result['errors'][0]}"
        raise HTTPException(status_code=502, detail=detail)

    conc.estado = "CERRADA"
    conc.factura_cliente_enviada = True
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)

    for a in archivos_leidos:
        db.add(FacturaArchivoCliente(
            conciliacion_id=conc.id,
            filename=a["filename"],
            content=a["content"],
            created_by=user.id,
        ))

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="envio_factura_cliente",
        valor_nuevo=f"destinatarios={', '.join(target_emails)}",
    )
    db.commit()
    db.refresh(conc)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Factura enviada al cliente",
        mensaje=f"La conciliacion '{conc.nombre}' fue facturada y cerrada con PDF enviado al cliente.",
        tipo="FACTURACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.get("/{conciliacion_id}/descargar-facturas")
def descargar_facturas_conciliacion(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    archivos = (
        db.query(FacturaArchivoCliente)
        .filter(FacturaArchivoCliente.conciliacion_id == conciliacion_id)
        .order_by(FacturaArchivoCliente.id.asc())
        .all()
    )
    if not archivos:
        raise HTTPException(status_code=404, detail="No hay facturas adjuntas para esta conciliacion")

    if len(archivos) == 1:
        a = archivos[0]
        return StreamingResponse(
            BytesIO(a.content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{a.filename}"'},
        )

    buf = BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for a in archivos:
            zf.writestr(a.filename, a.content)
    buf.seek(0)
    zip_name = f"facturas_conciliacion_{conciliacion_id}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_name}"'},
    )


@router.get("/{conciliacion_id}/descargar-excel")
def descargar_conciliacion_excel(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    items = (
        db.query(ConciliacionItem)
        .options(selectinload(ConciliacionItem.viaje).selectinload(Viaje.servicio))
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .order_by(ConciliacionItem.id.asc())
        .all()
    )
    avansat_prefetched = _prefetch_avansat_for_manifest_numbers_or_raise(
        db,
        [str(item.manifiesto_numero or "") for item in items],
    )

    placas = {str(item.placa or "").strip().upper() for item in items if item.placa}
    tipo_vehiculo_by_placa: dict[str, str] = {}
    if placas:
        vehiculos = (
            db.query(Vehiculo)
            .options(selectinload(Vehiculo.tipo))
            .filter(Vehiculo.placa.in_(list(placas)))
            .all()
        )
        tipo_vehiculo_by_placa = {
            str(v.placa or "").strip().upper(): (v.tipo.nombre if v.tipo else "")
            for v in vehiculos
        }

    excel_content = _build_conciliacion_excel(
        conc,
        items,
        user.rol,
        tipo_vehiculo_by_placa,
        avansat_prefetched,
    )
    filename = f"conciliacion_{conc.id}_resumen.xlsx"
    return StreamingResponse(
        BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{conciliacion_id}/cerrar", response_model=ConciliacionOut)
def cerrar_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede cerrar conciliacion")
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    if conc.estado != "APROBADA":
        raise HTTPException(status_code=400, detail="Solo conciliaciones aprobadas pueden cerrarse")
    operacion = db.get(Operacion, conc.operacion_id)
    conc.estado = "CERRADA"
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="cierre_conciliacion",
        valor_nuevo=payload.observacion or "cierre formal",
    )
    db.commit()
    db.refresh(conc)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA, UserRole.CLIENTE, UserRole.TERCERO])
    create_internal_notifications(
        db,
        recipients,
        titulo="Conciliacion cerrada",
        mensaje=f"La conciliacion '{conc.nombre}' fue cerrada formalmente.",
        tipo="CIERRE",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/comentarios", response_model=ComentarioOut)
def add_comment(
    payload: ComentarioCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol == UserRole.TERCERO:
        raise HTTPException(status_code=403, detail="Tercero no puede agregar comentarios")
    conc = db.get(Conciliacion, payload.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    if payload.item_id:
        item = db.get(ConciliacionItem, payload.item_id)
        if not item or item.conciliacion_id != payload.conciliacion_id:
            raise HTTPException(status_code=400, detail="Item invalido para la conciliacion")

    comment = Comentario(
        conciliacion_id=payload.conciliacion_id,
        item_id=payload.item_id,
        usuario_id=user.id,
        comentario=payload.comentario,
    )
    db.add(comment)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=payload.conciliacion_id,
        item_id=payload.item_id,
        campo="comentario",
        valor_nuevo=payload.comentario,
    )
    db.commit()
    db.refresh(comment)
    return comment


@router.get("/{conciliacion_id}/comentarios", response_model=list[ComentarioOut])
def get_comments(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    return (
        db.query(Comentario)
        .filter(Comentario.conciliacion_id == conciliacion_id)
        .order_by(Comentario.id.desc())
        .all()
    )


@router.get("/{conciliacion_id}/viajes-pendientes", response_model=list[dict])
def get_pending_viajes(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo conciliaciones en BORRADOR permiten adjuntar viajes")

    already_linked_viaje_ids = _existing_item_viaje_ids(db)

    viajes = (
        db.query(Viaje)
        .filter(
            Viaje.operacion_id == conc.operacion_id,
            Viaje.conciliacion_id.is_(None),
            Viaje.activo.is_(True),
        )
        .order_by(Viaje.fecha_servicio.asc(), Viaje.id.asc())
        .all()
    )

    if already_linked_viaje_ids:
        viajes = [v for v in viajes if v.id not in already_linked_viaje_ids]

    payload: list[dict] = []
    for viaje in viajes:
        out = ViajeOut.model_validate(viaje).model_dump()
        out["estado_conciliacion"] = _estado_conciliacion_viaje(viaje)
        if viaje.servicio:
            out["servicio_nombre"] = viaje.servicio.nombre
            out["servicio_codigo"] = viaje.servicio.codigo
        payload.append(out)

    return payload


@router.post("/{conciliacion_id}/adjuntar-viajes", response_model=list[ConciliacionItemOut])
def attach_pending_viajes(
    conciliacion_id: int,
    payload: AdjuntarViajesRequest,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede adjuntar viajes")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo conciliaciones en BORRADOR permiten adjuntar viajes")

    already_linked_viaje_ids = _existing_item_viaje_ids(db)

    viajes = (
        db.query(Viaje)
        .filter(
            Viaje.id.in_(payload.viaje_ids),
            Viaje.operacion_id == conc.operacion_id,
            Viaje.conciliacion_id.is_(None),
        )
        .all()
    )

    if already_linked_viaje_ids:
        viajes = [v for v in viajes if v.id not in already_linked_viaje_ids]

    if not viajes:
        raise HTTPException(status_code=400, detail="No hay viajes pendientes validos para adjuntar")

    created_items: list[ConciliacionItem] = []
    for viaje in viajes:
        tarifa_tercero, tarifa_cliente, rentabilidad = _default_viaje_item_financials(viaje)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            viaje_id=viaje.id,
            tipo=ItemTipo.VIAJE,
            fecha_servicio=viaje.fecha_servicio,
            origen=viaje.origen,
            destino=viaje.destino,
            placa=viaje.placa,
            conductor=viaje.conductor,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=tarifa_cliente,
            rentabilidad=rentabilidad,
            manifiesto_numero=viaje.manifiesto_numero,
            remesa=None,
            descripcion=viaje.descripcion,
            created_by=user.id,
            cargado_por=viaje.cargado_por,
        )
        estado_valor = getattr(conc.estado, "value", conc.estado)
        viaje.conciliado = _should_mark_conciliado(estado_valor)
        viaje.estado_conciliacion = str(estado_valor)
        viaje.conciliacion_id = conc.id
        db.add(item)
        log_change(
            db,
            usuario_id=user.id,
            conciliacion_id=conc.id,
            campo="viaje_adjuntado",
            valor_nuevo=f"viaje_id={viaje.id}",
        )
        created_items.append(item)

    _mark_borrador_dirty(conc)

    db.commit()
    for item in created_items:
        db.refresh(item)
    return created_items


@router.delete("/{conciliacion_id}/viajes/{viaje_id}")
def detach_viaje_from_conciliacion(
    conciliacion_id: int,
    viaje_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede quitar viajes")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo puedes quitar viajes cuando la conciliacion esta en BORRADOR")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    viaje = db.get(Viaje, viaje_id)
    if not viaje or viaje.conciliacion_id != conciliacion_id:
        raise HTTPException(status_code=404, detail="Viaje no encontrado en esta conciliacion")

    # Obtener IDs de los items a eliminar para limpiar historial primero.
    item_ids_to_delete = [
        row[0]
        for row in db.query(ConciliacionItem.id)
        .filter(
            ConciliacionItem.conciliacion_id == conciliacion_id,
            ConciliacionItem.tipo == ItemTipo.VIAJE,
            ConciliacionItem.viaje_id == viaje_id,
        )
        .all()
    ]

    if item_ids_to_delete:
        # Desvincula historial_cambios para evitar FK violation
        db.query(HistorialCambio).filter(
            HistorialCambio.item_id.in_(item_ids_to_delete)
        ).update({HistorialCambio.item_id: None}, synchronize_session=False)

        # Limpia todos los items VIAJE vinculados para ese viaje en esta conciliacion.
        db.query(ConciliacionItem).filter(
            ConciliacionItem.id.in_(item_ids_to_delete)
        ).delete(synchronize_session=False)

    viaje.conciliacion_id = None
    viaje.estado_conciliacion = None
    viaje.conciliado = False

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conciliacion_id,
        campo="viaje_desadjuntado",
        valor_nuevo=f"viaje_id={viaje_id}",
    )

    _mark_borrador_dirty(conc)

    db.commit()
    return {"ok": True}


@router.get("/{conciliacion_id}/historial", response_model=list[HistorialCambioOut])
def get_historial(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    return (
        db.query(HistorialCambio)
        .filter(HistorialCambio.conciliacion_id == conciliacion_id)
        .order_by(HistorialCambio.id.desc())
        .all()
    )


@router.get("/{conciliacion_id}/resumen-financiero", response_model=ResumenFinancieroOut)
def get_resumen_financiero(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    items = db.query(ConciliacionItem).filter(ConciliacionItem.conciliacion_id == conciliacion_id).all()
    total_tercero = sum(float(i.tarifa_tercero or 0) for i in items)
    total_cliente = sum(float(i.tarifa_cliente or 0) for i in items)
    total_rentabilidad_valor = total_cliente - total_tercero
    pct_vals = [float(i.rentabilidad) for i in items if i.rentabilidad is not None]
    pct_promedio = (sum(pct_vals) / len(pct_vals)) if pct_vals else 0

    if user.rol == UserRole.COINTRA:
        return {
            "total_tarifa_tercero": total_tercero,
            "total_tarifa_cliente": total_cliente,
            "total_rentabilidad_valor": total_rentabilidad_valor,
            "total_rentabilidad_pct_promedio": pct_promedio,
        }
    if user.rol == UserRole.CLIENTE:
        return {
            "total_tarifa_tercero": None,
            "total_tarifa_cliente": total_cliente,
            "total_rentabilidad_valor": None,
            "total_rentabilidad_pct_promedio": None,
        }
    return {
        "total_tarifa_tercero": total_tercero,
        "total_tarifa_cliente": None,
        "total_rentabilidad_valor": None,
        "total_rentabilidad_pct_promedio": None,
    }
