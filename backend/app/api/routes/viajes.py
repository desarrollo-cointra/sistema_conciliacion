from datetime import date, datetime, time
from io import BytesIO
import unicodedata

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user, is_cointra_admin
from app.db.session import get_db
from app.models.conciliacion import Conciliacion
from app.models.conciliacion_item import ConciliacionItem
from app.models.enums import UserRole
from app.models.enums import ItemTipo
from app.models.catalogo_tarifa import CatalogoTarifa
from app.models.operacion import Operacion
from app.models.servicio import Servicio
from app.models.usuario import Usuario
from app.models.usuario_operacion import usuario_operaciones_asignadas
from app.models.vehiculo import Vehiculo
from app.models.viaje import Viaje
from app.schemas.viaje import CargaMasivaFilaPreview, CargaMasivaResultado, ViajeCreate, ViajeOut, ViajeUpdate
from app.services.avansat_cache import resolve_avansat_from_cache_only
from app.services.pricing import calculate_tarifa_cliente

router = APIRouter(prefix="/viajes", tags=["viajes"])


def _estado_conciliacion_valor(viaje: Viaje) -> str | None:
    if viaje.estado_conciliacion:
        return viaje.estado_conciliacion
    if not viaje.conciliacion:
        return None
    estado = viaje.conciliacion.estado
    return getattr(estado, "value", estado)


def _expected_conciliado(viaje: Viaje) -> bool:
    estado = _estado_conciliacion_valor(viaje)
    if estado is None:
        return False
    return estado in {"APROBADA", "CERRADA"}


def _validate_user_access_operacion(user: Usuario, operacion: Operacion) -> None:
    if user.rol == UserRole.CLIENTE:
        is_assigned = any(op.id == operacion.id for op in user.operaciones_asignadas)
        if not is_assigned:
            raise HTTPException(status_code=403, detail="Operacion no disponible para este cliente")
    if user.rol == UserRole.TERCERO and user.tercero_id != operacion.tercero_id:
        raise HTTPException(status_code=403, detail="Operacion no disponible para este tercero")


def _ensure_cointra_admin(user: Usuario) -> None:
    if not is_cointra_admin(user):
        raise HTTPException(status_code=403, detail="Solo COINTRA_ADMIN puede editar o inactivar viajes")


def _ensure_viaje_mutable(viaje: Viaje, db: Session) -> None:
    """Evita alterar historicos cuando el viaje ya quedo dentro de una conciliacion en flujo avanzado."""
    if viaje.conciliacion_id is None:
        return

    conciliacion = viaje.conciliacion or db.get(Conciliacion, viaje.conciliacion_id)
    if not conciliacion:
        return

    estado = getattr(conciliacion.estado, "value", conciliacion.estado)
    if estado != "BORRADOR":
        raise HTTPException(
            status_code=400,
            detail=(
                "Este viaje pertenece a una conciliacion en estado "
                f"{estado}. No se puede modificar para proteger el historico liquidado/facturado."
            ),
        )


def _is_viaje_or_extra(servicio: Servicio | None) -> bool:
    if not servicio:
        return True
    return servicio.codigo in {"VIAJE", "VIAJE_ADICIONAL"}


def _is_hora_extra(servicio: Servicio | None) -> bool:
    if not servicio:
        return False
    return servicio.codigo == "HORA_EXTRA"


def _is_conductor_relevo(servicio: Servicio | None) -> bool:
    if not servicio:
        return False
    return servicio.codigo == "CONDUCTOR_RELEVO"


def _needs_origen_destino(servicio: Servicio | None) -> bool:
    """Retorna True si el servicio requiere origen y destino en formulario"""
    if not servicio:
        return True  # por defecto, si no hay servicio, se trata como viaje
    return servicio.requiere_origen_destino


def _calculate_horas_hasta_corte(hora_inicio: time, corte: time = time(6, 0)) -> float:
    start_minutes = hora_inicio.hour * 60 + hora_inicio.minute
    end_minutes = corte.hour * 60 + corte.minute
    total_minutes = (end_minutes - start_minutes) % (24 * 60)
    if total_minutes == 0:
        total_minutes = 24 * 60
    return round(total_minutes / 60, 2)


@router.post("", response_model=ViajeOut)
def create_viaje(
    payload: ViajeCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    # Crear viajes: COINTRA_ADMIN, COINTRA_USER, TERCERO
    if user.rol == UserRole.TERCERO:
        allowed = True
    elif user.rol == UserRole.COINTRA:
        # Cualquier sub_rol de Cointra puede crear viajes
        allowed = True
    else:
        allowed = False

    if not allowed:
        raise HTTPException(status_code=403, detail="No tiene permisos para crear viajes")

    operacion = db.get(Operacion, payload.operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    servicio = None
    if payload.servicio_id is not None:
        servicio = db.get(Servicio, payload.servicio_id)
        if not servicio or not servicio.activo:
            raise HTTPException(status_code=404, detail="Servicio no encontrado")

    is_conductor_relevo = _is_conductor_relevo(servicio)
    placa = "N/A" if is_conductor_relevo else payload.placa.strip().upper()
    
    origen = payload.origen.strip() if payload.origen and payload.origen.strip() else ""
    destino = payload.destino.strip() if payload.destino and payload.destino.strip() else ""

    # Validar que origen y destino estén presentes si el servicio así lo requiere
    if _needs_origen_destino(servicio):
        if not origen:
            raise HTTPException(status_code=400, detail="Origen es obligatorio para este servicio")
        if not destino:
            raise HTTPException(status_code=400, detail="Destino es obligatorio para este servicio")
    else:
        origen = origen or "N/A"
        destino = destino or "N/A"
    
    tarifa_tercero = float(payload.tarifa_tercero or 0)
    tarifa_cliente = float(payload.tarifa_cliente or 0)
    rentabilidad = None
    hora_inicio = payload.hora_inicio
    hora_fin = None
    horas_cantidad = None

    if is_conductor_relevo:
        hora_inicio = None
        hora_fin = None
        horas_cantidad = 1.0

        tarifa_manual_tercero = float(payload.tarifa_tercero or 0)
        tarifa_manual_cliente = float(payload.tarifa_cliente or 0)
        if tarifa_manual_tercero <= 0:
            raise HTTPException(
                status_code=400,
                detail="Debes ingresar tarifa manual para el servicio Conductor relevo.",
            )

        if tarifa_manual_cliente <= 0:
            tarifa_manual_cliente, rentabilidad = calculate_tarifa_cliente(tarifa_manual_tercero, operacion)
        else:
            if tarifa_manual_cliente > 0:
                rentabilidad = round((1 - (tarifa_manual_tercero / tarifa_manual_cliente)) * 100, 2)

        tarifa_tercero = round(tarifa_manual_tercero, 2)
        tarifa_cliente = round(float(tarifa_manual_cliente), 2)
    elif not _is_viaje_or_extra(servicio):
        vehiculo = db.query(Vehiculo).filter(Vehiculo.placa == placa, Vehiculo.activo.is_(True)).first()
        if not vehiculo:
            raise HTTPException(status_code=400, detail="La placa seleccionada no tiene vehiculo activo")

        tarifa_catalogo = (
            db.query(CatalogoTarifa)
            .filter(
                CatalogoTarifa.servicio_id == servicio.id,
                CatalogoTarifa.tipo_vehiculo_id == vehiculo.tipo_vehiculo_id,
                CatalogoTarifa.activo.is_(True),
            )
            .first()
        )
        factor = 1.0
        if _is_hora_extra(servicio):
            if not hora_inicio:
                raise HTTPException(status_code=400, detail="Hora inicio es obligatoria para servicio Hora Extra")
            hora_fin = time(6, 0)
            horas_cantidad = _calculate_horas_hasta_corte(hora_inicio, hora_fin)
            factor = float(horas_cantidad)
        else:
            hora_inicio = None
            hora_fin = None
            horas_cantidad = 1.0

        if tarifa_catalogo:
            tarifa_tercero = round(float(tarifa_catalogo.tarifa_tercero) * factor, 2)
            tarifa_cliente = round(float(tarifa_catalogo.tarifa_cliente) * factor, 2)
            rentabilidad = float(tarifa_catalogo.rentabilidad_pct)
        else:
            tarifa_manual_tercero = float(payload.tarifa_tercero or 0)
            tarifa_manual_cliente = float(payload.tarifa_cliente or 0)
            if tarifa_manual_tercero <= 0:
                raise HTTPException(
                    status_code=400,
                    detail="No hay tarifa parametrizada para ese servicio y tipo de vehiculo. Debes ingresar tarifa manual.",
                )

            if tarifa_manual_cliente <= 0:
                tarifa_manual_cliente, rentabilidad = calculate_tarifa_cliente(tarifa_manual_tercero, operacion)
            else:
                if tarifa_manual_cliente > 0:
                    rentabilidad = round((1 - (tarifa_manual_tercero / tarifa_manual_cliente)) * 100, 2)

            tarifa_tercero = round(tarifa_manual_tercero * factor, 2)
            tarifa_cliente = round(float(tarifa_manual_cliente) * factor, 2)

    if _is_viaje_or_extra(servicio) and tarifa_tercero and not tarifa_cliente:
        tarifa_cliente, rentabilidad = calculate_tarifa_cliente(float(tarifa_tercero), operacion)

    viaje = Viaje(
        operacion_id=payload.operacion_id,
        tercero_id=operacion.tercero_id,
        servicio_id=servicio.id if servicio else None,
        titulo=payload.titulo.strip(),
        fecha_servicio=payload.fecha_servicio,
        origen=origen,
        destino=destino,
        placa=placa,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
        horas_cantidad=horas_cantidad,
        conductor=None if is_conductor_relevo else payload.conductor,
        tarifa_tercero=tarifa_tercero,
        tarifa_cliente=tarifa_cliente,
        rentabilidad=rentabilidad,
        manifiesto_numero=payload.manifiesto_numero,
        descripcion=payload.descripcion,
        created_by=user.id,
        cargado_por=user.rol.value,
        estado_conciliacion=None,
        activo=True,
    )

    db.add(viaje)
    db.commit()
    db.refresh(viaje)
    return viaje


@router.get("", response_model=list[dict])
def list_viajes(
    operacion_id: int | None = None,
    only_pending: bool = False,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    query = (
        db.query(Viaje)
        .join(Operacion, Operacion.id == Viaje.operacion_id)
        .options(selectinload(Viaje.conciliacion), selectinload(Viaje.servicio))
    )
    if not is_cointra_admin(user):
        query = query.filter(Viaje.activo.is_(True))
    if user.rol == UserRole.CLIENTE:
        query = query.join(
            usuario_operaciones_asignadas,
            usuario_operaciones_asignadas.c.operacion_id == Operacion.id,
        ).filter(usuario_operaciones_asignadas.c.usuario_id == user.id)
    if user.rol == UserRole.TERCERO and user.tercero_id:
        query = query.filter(Operacion.tercero_id == user.tercero_id)
    if operacion_id:
        query = query.filter(Viaje.operacion_id == operacion_id)
    if only_pending:
        query = query.filter(Viaje.conciliado.is_(False))
    if fecha_desde:
        query = query.filter(Viaje.fecha_servicio >= fecha_desde)
    if fecha_hasta:
        query = query.filter(Viaje.fecha_servicio <= fecha_hasta)

    viajes = query.order_by(Viaje.id.desc()).all()

    changed = False

    # Repara enlaces historicos: item VIAJE existente sin viaje.conciliacion_id sincronizado.
    viaje_ids = [v.id for v in viajes]
    item_conciliacion_by_viaje_id: dict[int, int] = {}
    item_snapshot_by_viaje_id: dict[int, ConciliacionItem] = {}
    if viaje_ids:
        item_rows = (
            db.query(ConciliacionItem)
            .filter(
                ConciliacionItem.tipo == ItemTipo.VIAJE,
                ConciliacionItem.viaje_id.in_(viaje_ids),
                ConciliacionItem.viaje_id.is_not(None),
            )
            .order_by(ConciliacionItem.id.desc())
            .all()
        )
        for row in item_rows:
            viaje_id = row.viaje_id
            conciliacion_id = row.conciliacion_id
            if viaje_id is None:
                continue
            if viaje_id not in item_conciliacion_by_viaje_id:
                item_conciliacion_by_viaje_id[viaje_id] = conciliacion_id
                item_snapshot_by_viaje_id[viaje_id] = row

    conciliacion_ids_from_items = set(item_conciliacion_by_viaje_id.values())
    conciliaciones_map: dict[int, str] = {}
    if conciliacion_ids_from_items:
        conciliaciones = (
            db.query(Conciliacion.id, Conciliacion.estado)
            .filter(Conciliacion.id.in_(conciliacion_ids_from_items))
            .all()
        )
        conciliaciones_map = {cid: getattr(estado, "value", estado) for cid, estado in conciliaciones}

    for viaje in viajes:
        recovered_conciliacion_id = item_conciliacion_by_viaje_id.get(viaje.id)
        effective_conciliacion_id = viaje.conciliacion_id or recovered_conciliacion_id
        if viaje.conciliacion_id is None and recovered_conciliacion_id is not None:
            viaje.conciliacion_id = recovered_conciliacion_id
            changed = True

        estado_actual = _estado_conciliacion_valor(viaje)
        if estado_actual is None and effective_conciliacion_id is not None:
            estado_actual = conciliaciones_map.get(effective_conciliacion_id)

        esperado = _expected_conciliado(viaje)
        if estado_actual is not None:
            esperado = estado_actual in {"APROBADA", "CERRADA"}

        if viaje.conciliado != esperado:
            viaje.conciliado = esperado
            changed = True
        if viaje.estado_conciliacion != estado_actual:
            viaje.estado_conciliacion = estado_actual
            changed = True

    if changed:
        db.commit()

    payload: list[dict] = []
    for viaje in viajes:
        effective_estado = viaje.estado_conciliacion
        if user.rol == UserRole.CLIENTE and effective_estado not in {"EN_REVISION", "APROBADA", "CERRADA"}:
            continue

        out = ViajeOut.model_validate(viaje).model_dump()
        item_snapshot = item_snapshot_by_viaje_id.get(viaje.id)
        if item_snapshot:
            if item_snapshot.tarifa_tercero is not None:
                out["tarifa_tercero"] = float(item_snapshot.tarifa_tercero)
            if item_snapshot.tarifa_cliente is not None:
                out["tarifa_cliente"] = float(item_snapshot.tarifa_cliente)
            if item_snapshot.rentabilidad is not None:
                out["rentabilidad"] = float(item_snapshot.rentabilidad)

        if viaje.conciliacion_id is None and viaje.id in item_conciliacion_by_viaje_id:
            out["conciliacion_id"] = item_conciliacion_by_viaje_id[viaje.id]
        out["estado_conciliacion"] = effective_estado
        out["servicio_nombre"] = viaje.servicio.nombre if viaje.servicio else None
        out["servicio_codigo"] = viaje.servicio.codigo if viaje.servicio else None

        # Seguridad por API: cada rol solo recibe los valores financieros que le corresponden.
        if user.rol == UserRole.CLIENTE:
            out["tarifa_tercero"] = None
            out["rentabilidad"] = None
        elif user.rol == UserRole.TERCERO:
            out["tarifa_cliente"] = None
            out["rentabilidad"] = None

        payload.append(out)

    return payload


@router.patch("/{viaje_id}", response_model=ViajeOut)
def update_viaje(
    viaje_id: int,
    payload: ViajeUpdate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    viaje = db.get(Viaje, viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    _ensure_viaje_mutable(viaje, db)

    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No se enviaron cambios")

    if "tarifa_tercero" in data and data.get("tarifa_tercero") is not None:
        operacion = db.get(Operacion, viaje.operacion_id)
        tarifa_tercero = float(data["tarifa_tercero"])
        if "tarifa_cliente" not in data or data.get("tarifa_cliente") in (None, 0):
            tarifa_cliente, rentabilidad = calculate_tarifa_cliente(tarifa_tercero, operacion)
            data["tarifa_cliente"] = tarifa_cliente
            data["rentabilidad"] = rentabilidad

    for field, value in data.items():
        setattr(viaje, field, value)

    db.commit()
    db.refresh(viaje)
    return viaje


@router.delete("/{viaje_id}")
def deactivate_viaje(
    viaje_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    viaje = db.get(Viaje, viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    # COINTRA_ADMIN puede inactivar cualquier viaje sin restricción de estado
    viaje.activo = False
    db.commit()
    return {"ok": True}


@router.post("/{viaje_id}/reactivar")
def reactivate_viaje(
    viaje_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    viaje = db.get(Viaje, viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    _ensure_viaje_mutable(viaje, db)

    viaje.activo = True
    db.commit()
    return {"ok": True}


def _strip_accents(s: str) -> str:
    """Quita tildes y diacríticos para comparación robusta."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def _normalize_excel_header(h: str) -> str:
    """Normaliza un encabezado de columna Excel al nombre interno."""
    h = _strip_accents(h.lower().strip())
    if "tipo" in h and "servicio" in h:
        return "tipo_servicio"
    if "titul" in h:
        return "titulo"
    if "fecha" in h:
        return "fecha_servicio"
    if "placa" in h:
        return "placa"
    if "origen" in h:
        return "origen"
    if "destino" in h:
        return "destino"
    if "conductor" in h:
        return "conductor"
    if "tarifa" in h and "tercero" in h:
        return "tarifa_tercero"
    if "descripci" in h:
        return "descripcion"
    if "manifiesto" in h:
        return "manifiesto_numero"
    return h


def _normalize_manifiesto_bulk(value: object) -> str:
    """Normaliza número de manifiesto (igual que en avansat_cache)."""
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.endswith(".0"):
        integer_part = raw[:-2]
        if integer_part.isdigit():
            return integer_part
    return raw


def _normalize_placa_bulk(value: object) -> str:
    """Normaliza placa a alfanumérico mayúscula para comparación."""
    raw = str(value or "").strip().upper()
    return "".join(ch for ch in raw if ch.isalnum())


def _validate_manifiesto_bulk(db: Session, manifiesto: str, placa: str) -> str | None:
    """
    Valida el manifiesto para carga masiva.
    Retorna None si todo está correcto, o un string con el error encontrado.
    """
    resolved, _ = resolve_avansat_from_cache_only(db, [manifiesto])
    avansat = resolved.get(manifiesto) or {}
    if not avansat:
        return f"Manifiesto '{manifiesto}' no existe en la cache de Avansat"

    placa_avansat = _normalize_placa_bulk(avansat.get("placa_vehiculo") or "")
    placa_viaje = _normalize_placa_bulk(placa)
    if placa_avansat and placa_viaje and placa_avansat != placa_viaje:
        return (
            f"La placa del manifiesto ({placa_avansat}) "
            f"no coincide con la placa del viaje ({placa_viaje})"
        )

    existing = (
        db.query(Viaje.id)
        .filter(Viaje.manifiesto_numero == manifiesto)
        .first()
    )
    if existing:
        return f"El manifiesto '{manifiesto}' ya está asociado a otro viaje (#{existing[0]})"

    return None


def _get_servicios_viaje_map(db: Session) -> dict[str, Servicio]:
    """Retorna mapa codigo/nombre -> Servicio para VIAJE y VIAJE_ADICIONAL."""
    servicios = (
        db.query(Servicio)
        .filter(Servicio.codigo.in_(["VIAJE", "VIAJE_ADICIONAL"]), Servicio.activo.is_(True))
        .all()
    )
    result: dict[str, Servicio] = {}
    for s in servicios:
        result[s.codigo.upper()] = s
        result[s.nombre.upper()] = s
        # alias amigable "VIAJE ADICIONAL" (sin guion bajo)
        result[s.codigo.upper().replace("_", " ")] = s
    return result


@router.get("/plantilla-excel")
def descargar_plantilla_viajes():
    """Descarga una plantilla Excel lista para llenar con viajes masivos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Viajes"

    col_labels = [
        "tipo_servicio",
        "titulo",
        "fecha_servicio",
        "placa",
        "origen",
        "destino",
        "conductor",
        "tarifa_tercero",
        "descripcion",
        "manifiesto_numero",
    ]
    col_display = [
        "Tipo Servicio",
        "Título",
        "Fecha (YYYY-MM-DD)",
        "Placa",
        "Origen",
        "Destino",
        "Conductor (opcional)",
        "Tarifa Tercero",
        "Descripción (opcional)",
        "Manifiesto (opcional)",
    ]
    col_widths = [22, 30, 20, 14, 22, 22, 22, 18, 30, 22]

    header_fill = PatternFill(start_color="1F6B3A", end_color="1F6B3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    from openpyxl.utils import get_column_letter

    for col_num, (label, display, width) in enumerate(zip(col_labels, col_display, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=display)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Fila de ejemplo
    examples = [
        "VIAJE",
        "Urbano Ruta 1",
        "2026-05-15",
        "ABC123",
        "Ciudad A",
        "Ciudad B",
        "Juan Pérez",
        150000,
        "Servicio regular",
        "",
    ]
    for col_num, val in enumerate(examples, 1):
        ws.cell(row=2, column=col_num, value=val)

    # Instrucciones en hoja separada
    ws_info = wb.create_sheet("Instrucciones")
    ws_info["A1"] = "INSTRUCCIONES PARA CARGA MASIVA DE VIAJES"
    ws_info["A1"].font = Font(bold=True, size=13)
    instrucciones = [
        "",
        "1. En la hoja 'Viajes' complete una fila por cada servicio a cargar.",
        "2. La operación se selecciona en el sistema al momento de subir el archivo.",
        "3. Tipo Servicio: use exactamente 'VIAJE' o 'VIAJE_ADICIONAL'.",
        "4. Fecha: formato YYYY-MM-DD (ej: 2026-05-15).",
        "5. Placa: solo la placa del vehículo (ej: ABC123).",
        "6. Tarifa Tercero: número sin puntos ni comas (ej: 150000).",
        "7. Origen y Destino son obligatorios para tipo VIAJE y VIAJE_ADICIONAL.",
        "8. Conductor y Descripción son opcionales.",
        "9. No modifique los encabezados de la fila 1.",
        "10. Elimine la fila de ejemplo (fila 2) antes de cargar.",
        "11. Manifiesto: opcional. Si se indica, debe existir en Avansat y la placa debe coincidir.",
    ]
    for i, linea in enumerate(instrucciones, 2):
        ws_info[f"A{i}"] = linea
    ws_info.column_dimensions["A"].width = 70

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_viajes.xlsx"},
    )


@router.post("/carga-masiva/preview", response_model=list[CargaMasivaFilaPreview])
async def preview_carga_masiva_viajes(
    operacion_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Valida el archivo Excel sin guardar nada. Devuelve fila a fila el resultado."""
    if user.rol not in (UserRole.TERCERO, UserRole.COINTRA):
        raise HTTPException(status_code=403, detail="No tiene permisos para cargar viajes")

    operacion = db.get(Operacion, operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    content = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx)")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Archivo Excel vacío")

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    normalized = [_normalize_excel_header(h) for h in raw_headers]
    idx = {name: i for i, name in enumerate(normalized) if name}

    servicios_map = _get_servicios_viaje_map(db)
    seen_manifiestos: set[str] = set()  # control de duplicados en el mismo archivo

    def get_cell(row: tuple, key: str) -> str | None:
        if key not in idx:
            return None
        val = row[idx[key]]
        if val is None:
            return None
        return str(val).strip() or None

    result: list[CargaMasivaFilaPreview] = []

    for row_num, row in enumerate(rows[1:], start=2):
        # Extraer valores
        tipo_servicio_raw = get_cell(row, "tipo_servicio")
        titulo = get_cell(row, "titulo")
        fecha_raw = row[idx["fecha_servicio"]] if "fecha_servicio" in idx else None
        placa_raw = get_cell(row, "placa")
        origen = get_cell(row, "origen")
        destino = get_cell(row, "destino")
        conductor = get_cell(row, "conductor")
        tarifa_raw = row[idx["tarifa_tercero"]] if "tarifa_tercero" in idx else None
        descripcion = get_cell(row, "descripcion")

        placa = placa_raw.upper() if placa_raw else None
        manifiesto_raw = get_cell(row, "manifiesto_numero")
        manifiesto: str | None = _normalize_manifiesto_bulk(manifiesto_raw) if manifiesto_raw else None
        fecha_str: str | None = None
        if fecha_raw is not None:
            try:
                if isinstance(fecha_raw, datetime):
                    fecha_str = str(fecha_raw.date())
                elif isinstance(fecha_raw, date):
                    fecha_str = str(fecha_raw)
                else:
                    fecha_str = str(date.fromisoformat(str(fecha_raw).strip()))
            except Exception:
                fecha_str = str(fecha_raw).strip() or None

        # Convertir tarifa
        tarifa_num: float | None = None
        if tarifa_raw is not None:
            try:
                tarifa_num = float(tarifa_raw)
            except Exception:
                pass

        # Validar
        error: str | None = None
        valido = True
        servicio: Servicio | None = None

        if not tipo_servicio_raw:
            error = "Tipo Servicio es obligatorio. Use: VIAJE o VIAJE_ADICIONAL"
            valido = False
        else:
            lookup = tipo_servicio_raw.strip().upper().replace(" ", "_")
            servicio = servicios_map.get(lookup) or servicios_map.get(tipo_servicio_raw.strip().upper())
            if not servicio:
                error = f"Tipo Servicio '{tipo_servicio_raw}' no reconocido. Use: VIAJE o VIAJE_ADICIONAL"
                valido = False

        if valido and not titulo:
            error = "Título es obligatorio"
            valido = False

        if valido and not fecha_str:
            error = "Fecha es obligatoria (formato YYYY-MM-DD)"
            valido = False
        elif valido and fecha_str:
            try:
                date.fromisoformat(fecha_str)
            except ValueError:
                error = f"Fecha '{fecha_str}' no tiene formato válido (use YYYY-MM-DD)"
                valido = False

        if valido and not placa:
            error = "Placa es obligatoria"
            valido = False

        if valido and tarifa_num is None:
            error = "Tarifa Tercero es obligatoria y debe ser un número"
            valido = False
        elif valido and tarifa_num is not None and tarifa_num <= 0:
            error = "Tarifa Tercero debe ser mayor a 0"
            valido = False

        if valido and servicio:
            if servicio.requiere_origen_destino:
                if not origen:
                    error = "Origen es obligatorio para este tipo de servicio"
                    valido = False
                elif not destino:
                    error = "Destino es obligatorio para este tipo de servicio"
                    valido = False

        # Validar manifiesto si se proporcionó
        if valido and manifiesto:
            if manifiesto in seen_manifiestos:
                error = f"Manifiesto '{manifiesto}' aparece más de una vez en el archivo"
                valido = False
            else:
                seen_manifiestos.add(manifiesto)
                man_error = _validate_manifiesto_bulk(db, manifiesto, placa or "")
                if man_error:
                    error = man_error
                    valido = False

        result.append(
            CargaMasivaFilaPreview(
                fila=row_num,
                tipo_servicio=tipo_servicio_raw,
                titulo=titulo,
                fecha_servicio=fecha_str,
                placa=placa,
                origen=origen,
                destino=destino,
                conductor=conductor,
                tarifa_tercero=tarifa_num,
                descripcion=descripcion,
                manifiesto_numero=manifiesto,
                valido=valido,
                error=error,
            )
        )

    return result


@router.post("/carga-masiva", response_model=CargaMasivaResultado)
async def bulk_upload_viajes(
    operacion_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol not in (UserRole.TERCERO, UserRole.COINTRA):
        raise HTTPException(status_code=403, detail="No tiene permisos para cargar viajes")

    operacion = db.get(Operacion, operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    content = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx)")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Archivo Excel vacío")

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    normalized = [_normalize_excel_header(h) for h in raw_headers]
    idx = {name: i for i, name in enumerate(normalized) if name}

    servicios_map = _get_servicios_viaje_map(db)
    seen_manifiestos_upload: set[str] = set()  # control de duplicados en el mismo archivo

    def get_cell(row: tuple, key: str) -> str | None:
        if key not in idx:
            return None
        val = row[idx[key]]
        if val is None:
            return None
        return str(val).strip() or None

    errores: list[str] = []
    cargados = 0

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            tipo_servicio_raw = get_cell(row, "tipo_servicio")
            titulo = get_cell(row, "titulo")
            fecha_raw = row[idx["fecha_servicio"]] if "fecha_servicio" in idx else None
            placa_raw = get_cell(row, "placa")
            origen = get_cell(row, "origen") or ""
            destino = get_cell(row, "destino") or ""
            conductor = get_cell(row, "conductor")
            tarifa_raw = row[idx["tarifa_tercero"]] if "tarifa_tercero" in idx else None
            descripcion = get_cell(row, "descripcion")

            placa = placa_raw.upper() if placa_raw else ""
            manifiesto_raw = get_cell(row, "manifiesto_numero")
            manifiesto: str | None = _normalize_manifiesto_bulk(manifiesto_raw) if manifiesto_raw else None

            if not titulo:
                raise ValueError("Título es obligatorio")
            if fecha_raw is None:
                raise ValueError("Fecha es obligatoria")

            if isinstance(fecha_raw, datetime):
                fecha_servicio = fecha_raw.date()
            elif isinstance(fecha_raw, date):
                fecha_servicio = fecha_raw
            else:
                fecha_servicio = date.fromisoformat(str(fecha_raw).strip())

            if not placa:
                raise ValueError("Placa es obligatoria")
            if tarifa_raw is None:
                raise ValueError("Tarifa Tercero es obligatoria")

            tarifa_num = float(tarifa_raw)
            if tarifa_num <= 0:
                raise ValueError("Tarifa Tercero debe ser mayor a 0")

            # Resolver servicio
            servicio: Servicio | None = None
            if tipo_servicio_raw:
                lookup = tipo_servicio_raw.strip().upper().replace(" ", "_")
                servicio = servicios_map.get(lookup) or servicios_map.get(tipo_servicio_raw.strip().upper())

            if servicio and servicio.requiere_origen_destino:
                if not origen:
                    raise ValueError("Origen es obligatorio para este tipo de servicio")
                if not destino:
                    raise ValueError("Destino es obligatorio para este tipo de servicio")

            if not origen:
                origen = "N/A"
            if not destino:
                destino = "N/A"

            # Validar manifiesto si se proporcionó
            if manifiesto:
                if manifiesto in seen_manifiestos_upload:
                    raise ValueError(f"Manifiesto '{manifiesto}' aparece más de una vez en el archivo")
                man_error = _validate_manifiesto_bulk(db, manifiesto, placa)
                if man_error:
                    raise ValueError(man_error)
                seen_manifiestos_upload.add(manifiesto)

            tarifa_cliente, rentabilidad = calculate_tarifa_cliente(tarifa_num, operacion)

            viaje = Viaje(
                operacion_id=operacion_id,
                tercero_id=operacion.tercero_id,
                servicio_id=servicio.id if servicio else None,
                titulo=titulo,
                fecha_servicio=fecha_servicio,
                origen=origen,
                destino=destino,
                placa=placa,
                conductor=conductor or None,
                tarifa_tercero=round(tarifa_num, 2),
                tarifa_cliente=round(float(tarifa_cliente), 2),
                rentabilidad=rentabilidad,
                descripcion=descripcion,
                manifiesto_numero=manifiesto,
                created_by=user.id,
                cargado_por=user.rol.value,
                activo=True,
            )
            db.add(viaje)
            cargados += 1
        except Exception as exc:
            errores.append(f"Fila {row_num}: {exc}")

    db.commit()
    return CargaMasivaResultado(total_filas=max(0, len(rows) - 1), cargados=cargados, errores=errores)
